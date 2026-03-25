"""
recover_from_json.py
────────────────────────────────────────────────────────────
既存の dataset_*.json から xlsx / csv / json を再生成するスクリプト。

使い方:
  python recover_from_json.py
  python recover_from_json.py dataset_2026_03_21.json  ← ファイル指定も可

依存: pip install openpyxl
"""

import sys
import json
import csv
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── 設定 ────────────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).parent

CHECKBOX_COLUMNS = set([
    "Need available", "non need available", "both not available",
    "Agriculture", "Natural resources and conservation", "Architecture",
    "Area, ethnic, and gender studies", "Communication/journalism",
    "Communication technologies", "Computer and information sciences",
    "Personal and culinary services", "Education", "Engineering",
    "Engineering technologies", "Foreign languages, literatures, and linguistics",
    "Family and consumer sciences", "Law/legal studies", "English",
    "Liberal arts/general studies", "Library science", "Biological/life sciences",
    "Mathematics and statistics", "Military science and military technologies",
    "Interdisciplinary studies", "Parks and recreation",
    "Philosophy and religious studies", "Theology and religious vocations",
    "Physical sciences", "Science technologies", "Psychology",
    "Homeland Security, law enforcement, firefighting, and protective services",
    "Public administration and social services", "Social sciences",
    "Construction trades", "Mechanic and repair technologies",
    "Precision production", "Transportation and materials moving",
    "Visual and performing arts", "Health professions and related programs",
    "Business/marketing", "History",
    "スコア提出はあるがスコア条件はなし",
    "全額免除奨学金", "授業料免除奨学金", "Need-based Scholarship", "Need-Met 100%",
    "柳井正財団", "笹川平和財団", "グルーバンクロフト基金", "東進海外進学支援制度",
    "江副リクルート記念財団", "YouAreWelcomeHere Scholarship",
    "Stamps Scholar Program", "Laidlaw Scholars",
    "日本の全財団奨学生の進学した大学",
    "学部生のみ", "大学院生もいるがかなり少ない", "大学院生の割合が多い",
    "PhD課程あり", "研究型大学", "大学群・ラベリングのある大学",
    "日本人学生が10人以下の大学", "日本からの直行便がある",
    "空港から公共交通で90分圏内", "公共交通2時間圏内に大都市あり",
    "学生福祉（バス無料など）", "ウーバーや配車サービスあり", "24時間図書館あり",
])

COLUMNS = [
    "大学名", "国", "都市", "緯度", "経度", "URL", "気候",
    "学生数", "全体合格率",
    "男性受験者数", "女性受験者数", "男性合格者数", "女性合格者数",
    "男性入学者数", "女性入学者数",
    "全体出願者数", "全体合格者数", "全体入学者数",
    "留学生出願者数", "留学生合格者数", "留学生入学者数",
    "SAT提出割合", "ACT提出割合",
    "SAT Composite", "SAT EBRW", "SAT Math",
    "ACT Composite", "ACT Math", "ACT Science", "ACT Reading",
    "GPA", "区間の下限", "下限までの累計", "区間のパーセンテージ",
    "Priority Date", "Deferred Admission",
    "ED出願者数", "ED合格者数", "EA出願者数", "EA合格者数",
    "通貨定義",
    "Tuition", "文系平均", "理系平均", "現地語平均",
    "Required Fees", "Food and housing total", "Housing Only", "Food Only",
    "Books and supplies",
    "Need総数", "Need-Met", "Need全体平均額",
    "Merit全体数", "Merit平均額",
    "Need available", "non need available", "both not available",
    "Need留学生総数", "Need留学生平均額",
    "2-9", "10-19", "20-29", "30-39", "40-49", "50-99", "100+", "Total",
    "Agriculture", "Natural resources and conservation", "Architecture",
    "Area, ethnic, and gender studies", "Communication/journalism",
    "Communication technologies", "Computer and information sciences",
    "Personal and culinary services", "Education", "Engineering",
    "Engineering technologies", "Foreign languages, literatures, and linguistics",
    "Family and consumer sciences", "Law/legal studies", "English",
    "Liberal arts/general studies", "Library science", "Biological/life sciences",
    "Mathematics and statistics", "Military science and military technologies",
    "Interdisciplinary studies", "Parks and recreation",
    "Philosophy and religious studies", "Theology and religious vocations",
    "Physical sciences", "Science technologies", "Psychology",
    "Homeland Security, law enforcement, firefighting, and protective services",
    "Public administration and social services", "Social sciences",
    "Construction trades", "Mechanic and repair technologies",
    "Precision production", "Transportation and materials moving",
    "Visual and performing arts", "Health professions and related programs",
    "Business/marketing", "History",
    "TOEFL 総合", "TOEFL Reading", "TOEFL Listening", "TOEFL Writing", "TOEFL Speaking",
    "IELTS 総合", "IELTS Reading", "IELTS Listening", "IELTS Writing", "IELTS Speaking",
    "Duolingo 総合", "Duolingo Literacy", "Duolingo Comprehension",
    "Duolingo Conversation", "Duolingo Production",
    "スコア提出はあるがスコア条件はなし",
    "全額免除奨学金", "授業料免除奨学金", "Need-based Scholarship", "Need-Met 100%",
    "柳井正財団", "笹川平和財団", "グルーバンクロフト基金", "東進海外進学支援制度",
    "江副リクルート記念財団", "YouAreWelcomeHere Scholarship",
    "Stamps Scholar Program", "Laidlaw Scholars",
    "日本の全財団奨学生の進学した大学",
    "奨学金が反映される出願締め切り月", "大学入試最終出願締め切り月",
    "学部生のみ", "大学院生もいるがかなり少ない", "大学院生の割合が多い",
    "PhD課程あり", "研究型大学", "大学群・ラベリングのある大学",
    "日本人学生が10人以下の大学", "日本からの直行便がある",
    "空港から公共交通で90分圏内", "公共交通2時間圏内に大都市あり",
    "学生福祉（バス無料など）", "ウーバーや配車サービスあり", "24時間図書館あり",
]

SECTION_COLORS = {
    "大学名":       "1F4E79",
    "学生数":       "1F497D",
    "通貨定義":     "375623",
    "Need総数":     "7B3F00",
    "2-9":          "404040",
    "Agriculture":  "4B0082",
    "TOEFL 総合":   "006400",
    "全額免除奨学金":"8B0000",
    "学部生のみ":   "00008B",
}

def get_header_color(col_name, col_idx, columns):
    current_color = "1F4E79"
    for name, color in SECTION_COLORS.items():
        try:
            section_start = columns.index(name)
        except ValueError:
            continue
        if col_idx - 1 >= section_start:
            current_color = color
    return current_color


def get_output_stem(label=""):
    today = datetime.now().strftime("%Y_%m_%d")
    suffix = f"_{label}" if label else ""
    base = OUTPUT_DIR / f"dataset_{today}{suffix}"
    if not base.with_suffix(".xlsx").exists() and not base.with_suffix(".csv").exists():
        return base
    i = 1
    while True:
        candidate = OUTPUT_DIR / f"dataset_{today}{suffix}({i})"
        if not candidate.with_suffix(".xlsx").exists() and not candidate.with_suffix(".csv").exists():
            return candidate
        i += 1


def build_excel(universities, output_path):
    print(f"  Excel 作成中... ({len(universities)} 大学 × {len(COLUMNS)} カラム)")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "大学データ"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ヘッダー行
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        color = get_header_color(col_name, col_idx, COLUMNS)
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = PatternFill("solid", start_color=color, end_color=color)
        cell.alignment = header_align
    ws.row_dimensions[1].height = 45

    data_font = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
    cb_fill   = PatternFill("solid", start_color="F0F8E8", end_color="F0F8E8")

    # チェックボックス列インデックス（1-based）
    cb_col_indices = {
        col_idx for col_idx, col_name in enumerate(COLUMNS, start=1)
        if col_name in CHECKBOX_COLUMNS
    }

    # ─────────────────────────────────────────────────────────
    # ★ 修正ポイント：DataValidation はセルごとではなく
    #   列ごとにまとめて範囲指定して登録（大規模データで落ちるバグを修正）
    # ─────────────────────────────────────────────────────────
    max_row = len(universities) + 1  # ヘッダー込み

    # データ行
    for row_idx, uni in enumerate(universities, start=2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = uni.get(col_name, None)
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx in cb_col_indices:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = cb_fill
                if value is None:
                    cell.value = "FALSE"
            else:
                cell.alignment = Alignment(
                    horizontal="left" if col_name in ("大学名", "URL") else "center",
                    vertical="center",
                )
                if is_alt:
                    cell.fill = alt_fill

    # DataValidation をチェックボックス列ごとに1つの範囲として追加
    for cb_col_idx in cb_col_indices:
        letter = get_column_letter(cb_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
            sqref=f"{letter}2:{letter}{max_row}",  # セル単位でなく列範囲で一括登録
        )
        dv.showDropDown = False
        ws.add_data_validation(dv)

    # 列幅
    col_widths = {
        "大学名": 42, "国": 16, "都市": 18, "URL": 52,
        "緯度": 13, "経度": 13, "気候": 14, "通貨定義": 12,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 13)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    print(f"  保存中: {output_path.name} ...", end=" ", flush=True)
    wb.save(output_path)
    print(f"✅ 完了")
    print(f"  ✅ Excel: {output_path.name}  ({len(universities)}大学 / {len(COLUMNS)}カラム)")


def build_csv(universities, output_path):
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for uni in universities:
            row = {}
            for col in COLUMNS:
                val = uni.get(col, "")
                if col in CHECKBOX_COLUMNS and (val is None or val == ""):
                    val = "FALSE"
                row[col] = val
            writer.writerow(row)
    print(f"  ✅ CSV:   {output_path.name}")


def main():
    # JSON ファイルを特定
    if len(sys.argv) >= 2:
        json_path = Path(sys.argv[1])
        if not json_path.is_absolute():
            json_path = OUTPUT_DIR / json_path
    else:
        # 引数なし → 最新の dataset_*.json を自動選択
        candidates = sorted(OUTPUT_DIR.glob("dataset_*.json"), reverse=True)
        if not candidates:
            print("[ERROR] dataset_*.json が見つかりません。")
            print("        python recover_from_json.py <ファイル名> で指定してください。")
            return
        json_path = candidates[0]

    if not json_path.exists():
        print(f"[ERROR] ファイルが見つかりません: {json_path}")
        return

    print("=" * 55)
    print("  JSON リカバリスクリプト")
    print("=" * 55)
    print(f"\n読み込み: {json_path.name}")

    with open(json_path, encoding="utf-8") as f:
        universities = json.load(f)

    print(f"  → {len(universities)} 大学のデータを読み込みました")

    stem      = get_output_stem("recovered")
    xlsx_path = stem.with_suffix(".xlsx")
    csv_path  = stem.with_suffix(".csv")
    out_json  = stem.with_suffix(".json")

    print(f"\n[STEP 1] Excel 出力")
    build_excel(universities, xlsx_path)

    print(f"\n[STEP 2] CSV 出力")
    build_csv(universities, csv_path)

    print(f"\n[STEP 3] JSON コピー")
    import shutil
    shutil.copy2(json_path, out_json)
    print(f"  ✅ JSON:  {out_json.name}")

    print(f"\n完了！ 出力先: {OUTPUT_DIR.resolve()}")
    print(f"  {xlsx_path.name}")
    print(f"  {csv_path.name}")
    print(f"  {out_json.name}")


if __name__ == "__main__":
    main()
