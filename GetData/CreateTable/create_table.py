"""
University Dataset Automation Script v2
========================================
使い方:
  pip install requests beautifulsoup4 openpyxl
  python university_dataset_v2.py

出力:
  dataset_YYYY_MM_DD.xlsx  ← フリーズ・チェックボックス・色分け付き
  dataset_YYYY_MM_DD.csv   ← 同内容のCSV（チェックボックスは TRUE/FALSE）
  ※ 既存ファイルがあれば (1)(2)... と連番

出力先: スクリプトと同じフォルダ
  例: C:/Users/user/OneDrive/Desktop/willabroadjapan/GetData/CreateTable/
"""

import re
import time
import json
import csv
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ────────────────────────────────────────────────────────────
# ★ 設定エリア ★
# ────────────────────────────────────────────────────────────

COUNTRIES = {
    "アイスランド":             "https://www.unirank.org/is/a-z/",
    "アイルランド":             "https://www.unirank.org/ie/a-z/",
    "アゼルバイジャン":         "https://www.unirank.org/az/a-z/",
    "アメリカ":                 "https://www.unirank.org/us/a-z/",
    "アメリカンバージンアイランド": "https://www.unirank.org/vi/ranking/",
    "アラブ首長国連邦":         "https://www.unirank.org/ae/a-z/",
    "アルゼンチン":             "https://www.unirank.org/ar/a-z/",
    "アルバニア":               "https://www.unirank.org/al/a-z/",
    "アルメニア":               "https://www.unirank.org/am/a-z/",
    "アンドラ":                 "https://www.unirank.org/ad/ranking/",
    "イギリス":                 "https://www.unirank.org/gb/a-z/",
    "イスラエル":               "https://www.unirank.org/il/a-z/",
    "イタリア":                 "https://www.unirank.org/it/a-z/",
    "インド":                   "https://www.unirank.org/in/a-z/",
    "インドネシア":             "https://www.unirank.org/id/a-z/",
    "ウズベキスタン":           "https://www.unirank.org/uz/a-z/",
    "ウルグアイ":               "https://www.unirank.org/uy/a-z/",
    "エクアドル":               "https://www.unirank.org/ec/a-z/",
    "エジプト":                 "https://www.unirank.org/eg/a-z/",
    "エストニア":               "https://www.unirank.org/ee/a-z/",
    "オーストラリア":           "https://www.unirank.org/au/a-z/",
    "オーストリア":             "https://www.unirank.org/at/a-z/",
    "オマーン":                 "https://www.unirank.org/om/a-z/",
    "オランダ":                 "https://www.unirank.org/nl/a-z/",
    "カザフスタン":             "https://www.unirank.org/kz/a-z/",
    "カタール":                 "https://www.unirank.org/qa/a-z/",
    "カナダ":                   "https://www.unirank.org/ca/a-z/",
    "韓国":                     "https://www.unirank.org/kr/a-z/",
    "北マケドニア":             "https://www.unirank.org/mk/a-z/",
    "キプロス":                 "https://www.unirank.org/cy/a-z/",
    "キュラソー":               "https://www.unirank.org/cw/a-z/",
    "キューバ":                 "https://www.unirank.org/cu/a-z/",
    "ギリシャ":                 "https://www.unirank.org/gr/a-z/",
    "キルギス":                 "https://www.unirank.org/kg/a-z/",
    "グアテマラ":               "https://www.unirank.org/gt/a-z/",
    "グアドループ":             "https://www.unirank.org/gp/ranking/",
    "グアム":                   "https://www.unirank.org/gu/a-z/",
    "クウェート":               "https://www.unirank.org/kw/a-z/",
    "グリーンランド":           "https://www.unirank.org/gl/a-z/",
    "グレナダ":                 "https://www.unirank.org/gd/a-z/",
    "クロアチア":               "https://www.unirank.org/hr/a-z/",
    "コスタリカ":               "https://www.unirank.org/cr/a-z/",
    "コソボ":                   "https://www.unirank.org/xk/a-z/",
    "コロンビア":               "https://www.unirank.org/co/a-z/",
    "サウジアラビア":           "https://www.unirank.org/sa/a-z/",
    "サモア":                   "https://www.unirank.org/ws/a-z/",
    "サンマリノ":               "https://www.unirank.org/sm/ranking/",
    "ジョージア":               "https://www.unirank.org/ge/a-z/",
    "シンガポール":             "https://www.unirank.org/sg/a-z/",
    "スイス":                   "https://www.unirank.org/ch/a-z/",
    "スウェーデン":             "https://www.unirank.org/se/a-z/",
    "スペイン":                 "https://www.unirank.org/es/a-z/",
    "スリランカ":               "https://www.unirank.org/lk/a-z/",
    "スロバキア":               "https://www.unirank.org/sk/a-z/",
    "スロベニア":               "https://www.unirank.org/si/a-z/",
    "セルビア":                 "https://www.unirank.org/rs/",
    "タイ":                     "https://www.unirank.org/th/a-z/",
    "台湾":                     "https://www.unirank.org/tw/a-z/",
    "チェコ":                   "https://www.unirank.org/cz/a-z/",
    "中国":                     "https://www.unirank.org/cn/a-z/",
    "デンマーク":               "https://www.unirank.org/dk/a-z/",
    "ドイツ":                   "https://www.unirank.org/de/a-z/",
    "ドミニカ共和国":           "https://www.unirank.org/do/a-z/",
    "トルコ":                   "https://www.unirank.org/tr/a-z/",
    "ニューカレドニア":         "https://www.unirank.org/nc/ranking/",
    "ニュージーランド":         "https://www.unirank.org/nz/a-z/",
    "ネパール":                 "https://www.unirank.org/np/a-z/",
    "ノルウェー":               "https://www.unirank.org/no/a-z/",
    "バーレーン":               "https://www.unirank.org/bh/a-z/",
    "パナマ":                   "https://www.unirank.org/pa/a-z/",
    "パプアニューギニア":       "https://www.unirank.org/pg/a-z/",
    "パラグアイ":               "https://www.unirank.org/py/a-z/",
    "ハンガリー":               "https://www.unirank.org/hu/a-z/",
    "バングラデシュ":           "https://www.unirank.org/bd/a-z/",
    "フィジー":                 "https://www.unirank.org/fj/a-z/",
    "フィリピン":               "https://www.unirank.org/ph/a-z/",
    "フィンランド":             "https://www.unirank.org/fi/a-z/",
    "プエルトリコ":             "https://www.unirank.org/pr/a-z/",
    "ブータン":                 "https://www.unirank.org/bt/a-z/",
    "ブラジル":                 "https://www.unirank.org/br/a-z/",
    "フランス":                 "https://www.unirank.org/fr/a-z/",
    "ブルガリア":               "https://www.unirank.org/bg/a-z/",
    "ブルネイ":                 "https://www.unirank.org/bn/a-z/",
    "フレンチポリネシア":       "https://www.unirank.org/pf/ranking/",
    "ベトナム":                 "https://www.unirank.org/vn/a-z/",
    "ペルー":                   "https://www.unirank.org/pe/a-z/",
    "ベルギー":                 "https://www.unirank.org/be/a-z/",
    "ポーランド":               "https://www.unirank.org/pl/a-z/",
    "ボスニアヘルツェゴビナ":   "https://www.unirank.org/ba/a-z/",
    "ボリビア":                 "https://www.unirank.org/bo/a-z/",
    "ポルトガル":               "https://www.unirank.org/pt/a-z/",
    "香港":                     "https://www.unirank.org/hk/a-z/",
    "マカオ":                   "https://www.unirank.org/mo/a-z/",
    "マルタ":                   "https://www.unirank.org/mt/a-z/",
    "マレーシア":               "https://www.unirank.org/my/a-z/",
    "南アフリカ":               "https://www.unirank.org/za/a-z/",
    "メキシコ":                 "https://www.unirank.org/mx/a-z/",
    "モーリシャス":             "https://www.unirank.org/mu/a-z/",
    "モナコ":                   "https://www.unirank.org/mc/ranking/",
    "モルディブ":               "https://www.unirank.org/mv/ranking/",
    "モルドバ":                 "https://www.unirank.org/md/a-z/",
    "モロッコ":                 "https://www.unirank.org/ma/a-z/",
    "モンゴル":                 "https://www.unirank.org/mn/a-z/",
    "モンテネグロ":             "https://www.unirank.org/me/ranking/",
    "ラオス":                   "https://www.unirank.org/la/ranking/",
    "ラトビア":                 "https://www.unirank.org/lv/a-z/",
    "リトアニア":               "https://www.unirank.org/lt/a-z/",
    "ルーマニア":               "https://www.unirank.org/ro/a-z/",
    "ルクセンブルク":           "https://www.unirank.org/lu/a-z/",
    "レユニオン":               "https://www.unirank.org/re/a-z/",
    "ヨルダン":                 "https://www.unirank.org/jo/a-z/",
}

# スクリプトと同じフォルダに出力
OUTPUT_DIR = Path(__file__).parent

TEST_MODE = False       # True: サンプルデータで動作確認
GEOCODE_DELAY     = 3.0   # 通常リクエスト間隔（秒）
GEOCODE_RETRY_MAX = 5     # レート制限時の最大リトライ回数
GEOCODE_RETRY_WAIT = 60   # レート制限検知時の待機秒数（429受信 or 連続失敗）
PROGRESS_FILE        = OUTPUT_DIR / "geocode_progress.json"   # ジオコーディング進捗（中断再開用）
SCRAPE_PROGRESS_FILE = OUTPUT_DIR / "scrape_progress.json"    # スクレイピング進捗（国ごとに保存）
RANKINGS_CACHE_FILE  = OUTPUT_DIR / "rankings_cache.json"     # THE/QS ランキングキャッシュ（年1回更新）
INPUT_DATA_FILE      = OUTPUT_DIR / "input_data.json"         # 手動入力データ（InputDataWithPanel.py で生成）
URL_BACKUP_FILE      = OUTPUT_DIR / "university_urls.json"    # URL バックアップ（url_checker.py で生成）
RANKINGS_CACHE_MAX_DAYS = 330  # 約11ヶ月で自動再取得（毎年6月頃に新ランキング公開）
DATA_LISTING_DIR     = OUTPUT_DIR / "data_listing"            # ezo_art_YYYY.pdf 等の配置フォルダ

# ────────────────────────────────────────────────────────────
# 通貨マッピング（国名 → 通貨記号）
# ────────────────────────────────────────────────────────────

CURRENCY_MAP = {
    # 米ドル
    "アメリカ": "USD", "アメリカンバージンアイランド": "USD",
    "グアム": "USD", "プエルトリコ": "USD", "パナマ": "USD",
    "エクアドル": "USD", "エルサルバドル": "USD", "東ティモール": "USD",
    # オーストラリアドル
    "オーストラリア": "AUD",
    # カナダドル
    "カナダ": "CAD",
    # 韓国ウォン
    "韓国": "KRW",
    # シンガポールドル
    "シンガポール": "SGD",
    # 中国元
    "中国": "CNY", "マカオ": "CNY",
    # ニュージーランドドル
    "ニュージーランド": "NZD",
    # ポンド
    "イギリス": "GBP",
    # ユーロ
    # アイスランドはISKクローナ → 指定通貨リスト外なのでUSD（下のデフォルト処理で対応）
    "アイルランド": "EUR", "イタリア": "EUR", "ドイツ": "EUR",
    "フランス": "EUR", "スペイン": "EUR", "オランダ": "EUR",
    "ベルギー": "EUR", "ポルトガル": "EUR", "オーストリア": "EUR",
    "フィンランド": "EUR", "ギリシャ": "EUR", "スロベニア": "EUR",
    "スロバキア": "EUR", "エストニア": "EUR", "ラトビア": "EUR",
    "リトアニア": "EUR", "ルクセンブルク": "EUR", "マルタ": "EUR",
    "キプロス": "EUR", "モナコ": "EUR", "アンドラ": "EUR",
    "サンマリノ": "EUR", "バチカン": "EUR",
    "グアドループ": "EUR", "レユニオン": "EUR", "フレンチポリネシア": "EUR",
    "ニューカレドニア": "EUR",
    # リンギット
    "マレーシア": "MYR",
    # UAEディルハム
    "アラブ首長国連邦": "AED",
    # ※ アイスランドはISKだが指定通貨リスト外 → USD扱い
}

def get_currency(country_name: str) -> str:
    """国名から通貨コードを返す。指定通貨リスト外はUSD。"""
    return CURRENCY_MAP.get(country_name, "USD")

# ────────────────────────────────────────────────────────────
# 気候マッピング（国名 → フィルター値）
#   oceanic      = 西岸海洋性気候
#   mediterranean= 地中海性気候
#   humid_sub    = 温暖温潤気候
#   subarctic    = 亜寒帯・冷帯
#   steppe       = ステップ・砂漠気候
#   tropical     = 熱帯雨林・サバナ気候
# ────────────────────────────────────────────────────────────

CLIMATE_MAP = {
    # 亜寒帯・冷帯 ───────────────────────────────────────────
    "アイスランド":             "subarctic",
    "グリーンランド":           "subarctic",
    "フィンランド":             "subarctic",
    "スウェーデン":             "subarctic",
    "ノルウェー":               "subarctic",
    "カナダ":                   "subarctic",
    "ロシア":                   "subarctic",
    "モンゴル":                 "subarctic",
    "キルギス":                 "subarctic",

    # 西岸海洋性気候 ─────────────────────────────────────────
    "アイルランド":             "oceanic",
    "イギリス":                 "oceanic",
    "オランダ":                 "oceanic",
    "ベルギー":                 "oceanic",
    "ルクセンブルク":           "oceanic",
    "フランス":                 "oceanic",
    "ドイツ":                   "oceanic",
    "デンマーク":               "oceanic",
    "オーストリア":             "oceanic",
    "スイス":                   "oceanic",
    "チェコ":                   "oceanic",
    "スロバキア":               "oceanic",
    "スロベニア":               "oceanic",
    "クロアチア":               "oceanic",
    "ポーランド":               "oceanic",
    "ハンガリー":               "oceanic",
    "ルーマニア":               "oceanic",
    "ブルガリア":               "oceanic",
    "セルビア":                 "oceanic",
    "ボスニアヘルツェゴビナ":   "oceanic",
    "モンテネグロ":             "oceanic",
    "北マケドニア":             "oceanic",
    "コソボ":                   "oceanic",
    "アルバニア":               "oceanic",
    "ニュージーランド":         "oceanic",
    "アンドラ":                 "oceanic",

    # 地中海性気候 ───────────────────────────────────────────
    "スペイン":                 "mediterranean",
    "ポルトガル":               "mediterranean",
    "イタリア":                 "mediterranean",
    "ギリシャ":                 "mediterranean",
    "キプロス":                 "mediterranean",
    "マルタ":                   "mediterranean",
    "モナコ":                   "mediterranean",
    "サンマリノ":               "mediterranean",
    "クロアチア":               "mediterranean",
    "イスラエル":               "mediterranean",
    "レユニオン":               "mediterranean",
    "南アフリカ":               "mediterranean",   # ケープタウン周辺

    # 温暖温潤気候 ───────────────────────────────────────────
    "アメリカ":                 "humid_sub",
    "アメリカンバージンアイランド": "humid_sub",
    "プエルトリコ":             "humid_sub",
    "グアム":                   "humid_sub",
    "日本":                     "humid_sub",
    "韓国":                     "humid_sub",
    "中国":                     "humid_sub",
    "台湾":                     "humid_sub",
    "香港":                     "humid_sub",
    "マカオ":                   "humid_sub",
    "アルゼンチン":             "humid_sub",
    "ウルグアイ":               "humid_sub",
    "ブラジル":                 "humid_sub",  # 南部は温暖温潤
    "チリ":                     "humid_sub",
    "パラグアイ":               "humid_sub",
    "ジョージア":               "humid_sub",
    "アルメニア":               "humid_sub",
    "アゼルバイジャン":         "humid_sub",
    "ウズベキスタン":           "humid_sub",
    "エストニア":               "humid_sub",
    "ラトビア":                 "humid_sub",
    "リトアニア":               "humid_sub",
    "モルドバ":                 "humid_sub",
    "ウクライナ":               "humid_sub",
    "ベラルーシ":               "humid_sub",

    # ステップ・砂漠気候 ─────────────────────────────────────
    "アラブ首長国連邦":         "steppe",
    "サウジアラビア":           "steppe",
    "カタール":                 "steppe",
    "クウェート":               "steppe",
    "バーレーン":               "steppe",
    "オマーン":                 "steppe",
    "ヨルダン":                 "steppe",
    "エジプト":                 "steppe",
    "モロッコ":                 "steppe",
    "カザフスタン":             "steppe",
    "モルドバ":                 "steppe",

    # 熱帯雨林・サバナ気候 ───────────────────────────────────
    "インド":                   "tropical",
    "インドネシア":             "tropical",
    "マレーシア":               "tropical",
    "シンガポール":             "tropical",
    "タイ":                     "tropical",
    "ベトナム":                 "tropical",
    "フィリピン":               "tropical",
    "バングラデシュ":           "tropical",
    "スリランカ":               "tropical",
    "ネパール":                 "tropical",
    "ブータン":                 "tropical",
    "ブルネイ":                 "tropical",
    "ラオス":                   "tropical",
    "ミャンマー":               "tropical",
    "カンボジア":               "tropical",
    "フィジー":                 "tropical",
    "サモア":                   "tropical",
    "パプアニューギニア":       "tropical",
    "フレンチポリネシア":       "tropical",
    "ニューカレドニア":         "tropical",
    "グレナダ":                 "tropical",
    "キュラソー":               "tropical",
    "グアドループ":             "tropical",
    "コスタリカ":               "tropical",
    "グアテマラ":               "tropical",
    "パナマ":                   "tropical",
    "キューバ":                 "tropical",
    "ドミニカ共和国":           "tropical",
    "コロンビア":               "tropical",
    "エクアドル":               "tropical",
    "ペルー":                   "tropical",
    "ボリビア":                 "tropical",
    "モーリシャス":             "tropical",
    "モルディブ":               "tropical",
    "メキシコ":                 "tropical",   # 南部は熱帯
    "オーストラリア":           "tropical",   # 北部は熱帯（大学はシドニー等にも多いが代表値）
}

def get_climate(country_name: str) -> str:
    """国名から気候コードを返す。未定義の場合は空文字。"""
    return CLIMATE_MAP.get(country_name, "")

# ────────────────────────────────────────────────────────────
# チェックボックス対象カラム
# ────────────────────────────────────────────────────────────

CHECKBOX_COLUMNS = set([
    # 奨学金留学生チェック
    "Need available", "non need available", "both not available",
    # 学問分野
    "Agriculture", "Natural resources and conservation", "Architecture",
    "Area, ethnic, and gender studies", "Communication/journalism",
    "Communication technologies", "Computer and information sciences",
    "Personal and culinary services", "Education", "Engineering",
    "Engineering technologies", "Foreign languages, literatures, and linguistics",
    "Family and consumer sciences", "Law/legal studies", "English",
    "Liberal arts/general studies", "Library science", "Biological/life sciences",
    "Mathematics and statistics", "Military science and military technologies",
    "Interdisciplinary studies", "Parks and recreation",
    "Philosophy and religious studies", "Theology and religious vocations",
    "Physical sciences", "Science technologies", "Psychology",
    "Homeland Security, law enforcement, firefighting, and protective services",
    "Public administration and social services", "Social sciences",
    "Construction trades", "Mechanic and repair technologies",
    "Precision production", "Transportation and materials moving",
    "Visual and performing arts", "Health professions and related programs",
    "Business/marketing", "History",
    # 語学条件（チェックボックス）
    "スコア提出はあるがスコア条件はなし",
    # 奨学金種別
    "全額免除奨学金", "授業料免除奨学金", "Need-based Scholarship", "Need-Met 100%",
    "柳井正財団", "笹川平和財団", "グルーバンクロフト基金", "東進海外進学支援制度",
    "江副リクルート記念財団", "YouAreWelcomeHere Scholarship",
    "Stamps Scholar Program", "Laidlaw Scholars",
    "日本の全財団奨学生の進学した大学",
    # 大学特性
    "学部生のみ", "大学院生もいるがかなり少ない", "大学院生の割合が多い",
    "PhD課程あり", "研究型大学", "大学群・ラベリングのある大学",
    "日本人学生が10人以下の大学", "日本からの直行便がある",
    "空港から公共交通で90分圏内", "公共交通2時間圏内に大都市あり",
    "学生福祉（バス無料など）", "ウーバーや配車サービスあり", "24時間図書館あり",
])

# ────────────────────────────────────────────────────────────
# 全カラム定義
# ────────────────────────────────────────────────────────────

COLUMNS = [
    "大学名", "国", "都市", "緯度", "経度", "URL", "気候",
    # CDS
    "学生数", "全体合格率",
    "男性受験者数", "女性受験者数", "男性合格者数", "女性合格者数",
    "男性入学者数", "女性入学者数",
    "全体出願者数", "全体合格者数", "全体入学者数",
    "留学生出願者数", "留学生合格者数", "留学生入学者数",
    "SAT提出割合", "ACT提出割合",
    "SAT Composite", "SAT EBRW", "SAT Math",
    "ACT Composite", "ACT Math", "ACT Science", "ACT Reading",
    "GPA", "区間の下限", "下限までの累計", "区間のパーセンテージ",
    "Priority Date", "Deferred Admission",
    "ED出願者数", "ED合格者数", "EA出願者数", "EA合格者数",
    # 費用（通貨定義を費用カラム群の直前に追加）
    "通貨定義",
    "Tuition", "文系平均", "理系平均", "現地語平均",
    "Required Fees", "Food and housing total", "Housing Only", "Food Only",
    "Books and supplies",
    # 奨学金
    "Need総数", "Need-Met", "Need全体平均額",
    "Merit全体数", "Merit平均額",
    "Need available", "non need available", "both not available",
    "Need留学生総数", "Need留学生平均額",
    # クラスサイズ
    "2-9", "10-19", "20-29", "30-39", "40-49", "50-99", "100+", "Total",
    # 学問分野
    "Agriculture", "Natural resources and conservation", "Architecture",
    "Area, ethnic, and gender studies", "Communication/journalism",
    "Communication technologies", "Computer and information sciences",
    "Personal and culinary services", "Education", "Engineering",
    "Engineering technologies", "Foreign languages, literatures, and linguistics",
    "Family and consumer sciences", "Law/legal studies", "English",
    "Liberal arts/general studies", "Library science", "Biological/life sciences",
    "Mathematics and statistics", "Military science and military technologies",
    "Interdisciplinary studies", "Parks and recreation",
    "Philosophy and religious studies", "Theology and religious vocations",
    "Physical sciences", "Science technologies", "Psychology",
    "Homeland Security, law enforcement, firefighting, and protective services",
    "Public administration and social services", "Social sciences",
    "Construction trades", "Mechanic and repair technologies",
    "Precision production", "Transportation and materials moving",
    "Visual and performing arts", "Health professions and related programs",
    "Business/marketing", "History",
    # 語学条件（スコア数値）
    "TOEFL 総合", "TOEFL Reading", "TOEFL Listening", "TOEFL Writing", "TOEFL Speaking",
    "IELTS 総合", "IELTS Reading", "IELTS Listening", "IELTS Writing", "IELTS Speaking",
    "Duolingo 総合", "Duolingo Literacy", "Duolingo Comprehension",
    "Duolingo Conversation", "Duolingo Production",
    # 語学条件（チェックボックス）
    "スコア提出はあるがスコア条件はなし",
    # ランキング（THE / QS 総合・分野別・芸術分野）
    "THE総合", "QS総合", "THE分野最高位", "QS分野最高位",
    "THE芸術順位", "QS芸術順位",
    # 奨学金種別チェック
    "全額免除奨学金", "授業料免除奨学金", "Need-based Scholarship", "Need-Met 100%",
    "柳井正財団", "笹川平和財団", "グルーバンクロフト基金", "東進海外進学支援制度",
    "江副リクルート記念財団", "YouAreWelcomeHere Scholarship",
    "Stamps Scholar Program", "Laidlaw Scholars",
    "日本の全財団奨学生の進学した大学",
    "奨学金が反映される出願締め切り月", "大学入試最終出願締め切り月",
    # 大学特性チェック
    "学部生のみ", "大学院生もいるがかなり少ない", "大学院生の割合が多い",
    "PhD課程あり", "研究型大学", "大学群・ラベリングのある大学",
    "日本人学生が10人以下の大学", "日本からの直行便がある",
    "空港から公共交通で90分圏内", "公共交通2時間圏内に大都市あり",
    "学生福祉（バス無料など）", "ウーバーや配車サービスあり", "24時間図書館あり",
]

HEADERS_HTTP = {"User-Agent": "Mozilla/5.0 (compatible; UniversityDatasetBot/1.0)"}

# ────────────────────────────────────────────────────────────
# ファイル名生成（連番）
# ────────────────────────────────────────────────────────────

def get_output_stem() -> Path:
    """重複しないファイル名のステム(拡張子なし)を返す。"""
    today = datetime.now().strftime("%Y_%m_%d")
    base = OUTPUT_DIR / f"dataset_{today}"
    if not base.with_suffix(".xlsx").exists() and not base.with_suffix(".csv").exists():
        return base
    i = 1
    while True:
        candidate = OUTPUT_DIR / f"dataset_{today}({i})"
        if not candidate.with_suffix(".xlsx").exists() and not candidate.with_suffix(".csv").exists():
            return candidate
        i += 1

# ────────────────────────────────────────────────────────────
# スクレイピング
# ────────────────────────────────────────────────────────────

def scrape_unirank(country_name: str, url: str) -> list:
    try:
        resp = requests.get(url, headers=HEADERS_HTTP, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [ERROR] {country_name}: {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    universities = []
    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue
        a_tag = tds[1].find("a")
        if not a_tag:
            continue
        name = a_tag.get_text(strip=True)
        href = a_tag.get("href", "")
        full_url = f"https://www.unirank.org{href}" if href.startswith("/") else href
        city = tds[2].get_text(strip=True)
        if name:
            universities.append({"大学名": name, "国": country_name, "都市": city, "URL": ""})
    print(f"  → {country_name}: {len(universities)}大学")
    return universities

# ────────────────────────────────────────────────────────────
# ジオコーディング（Nominatim）
# ────────────────────────────────────────────────────────────

def geocode(name: str, city: str, country: str) -> tuple:
    """Nominatim でジオコーディング。レート制限時はバックオフリトライ。"""
    base_url = "https://nominatim.openstreetmap.org/search"
    queries = [name, f"{name}, {city}, {country}"]

    for query in queries:
        for attempt in range(GEOCODE_RETRY_MAX):
            try:
                resp = requests.get(
                    base_url,
                    params={"q": query, "format": "json", "limit": 1},
                    headers={"User-Agent": "UniversityDatasetBot/1.0 (study-abroad-db)"},
                    timeout=15,
                )
                if resp.status_code == 429:
                    wait = GEOCODE_RETRY_WAIT * (attempt + 1)
                    print(f"\n    [429 Rate Limit] {wait}秒待機中...", flush=True)
                    time.sleep(wait)
                    continue
                data = resp.json()
                if data:
                    time.sleep(GEOCODE_DELAY)
                    return float(data[0]["lat"]), float(data[0]["lon"])
                break  # 結果なし → 次のクエリへ
            except Exception:
                wait = GEOCODE_RETRY_WAIT * (attempt + 1)
                print(f"\n    [ERROR] リトライ{attempt+1}/{GEOCODE_RETRY_MAX} {wait}秒待機...", flush=True)
                time.sleep(wait)
        time.sleep(GEOCODE_DELAY)
    return None, None

# ────────────────────────────────────────────────────────────
# テスト用サンプルデータ
# ────────────────────────────────────────────────────────────

def get_sample_data() -> list:
    return [
        {"大学名": "University of Iceland",           "国": "アイスランド", "都市": "Reykjavik", "緯度": 64.1391, "経度": -21.9525, "URL": ""},
        {"大学名": "Reykjavik University",             "国": "アイスランド", "都市": "Reykjavik", "緯度": 64.1178, "経度": -21.9305, "URL": ""},
        {"大学名": "Agricultural University of Iceland","国": "アイスランド", "都市": "Borgarnes",  "緯度": 64.5375, "経度": -21.9188, "URL": ""},
        {"大学名": "Trinity College Dublin",           "国": "アイルランド", "都市": "Dublin",    "緯度": 53.3438, "経度": -6.2546,  "URL": ""},
        {"大学名": "University College Dublin",        "国": "アイルランド", "都市": "Dublin",    "緯度": 53.3065, "経度": -6.2199,  "URL": ""},
        {"大学名": "University College Cork",          "国": "アイルランド", "都市": "Cork",      "緯度": 51.8936, "経度": -8.4919,  "URL": ""},
        {"大学名": "University of Galway",             "国": "アイルランド", "都市": "Galway",    "緯度": 53.2779, "経度": -9.0597,  "URL": ""},
        {"大学名": "Dublin City University",           "国": "アイルランド", "都市": "Dublin",    "緯度": 53.3862, "経度": -6.2575,  "URL": ""},
    ]

# ────────────────────────────────────────────────────────────
# 同一国内の重複大学名に (1)(2)... を付加
# ────────────────────────────────────────────────────────────

def deduplicate_names(universities: list):
    """同一国内で同名の大学が複数ある場合に (1)(2)... を付加する。"""
    # (国, 大学名) の出現数をカウント
    key_count: dict = {}
    for uni in universities:
        k = (uni["国"], uni["大学名"])
        key_count[k] = key_count.get(k, 0) + 1

    # 2件以上ある場合のみ連番付加
    key_seen: dict = {}
    for uni in universities:
        k = (uni["国"], uni["大学名"])
        if key_count[k] > 1:
            key_seen[k] = key_seen.get(k, 0) + 1
            uni["大学名"] = f"{uni['大学名']}({key_seen[k]})"

    dupes = sum(1 for v in key_count.values() if v > 1)
    if dupes:
        print(f"  → 重複大学名を {dupes} 件検出し、連番を付加しました")


# ────────────────────────────────────────────────────────────
# 通貨を自動補完
# ────────────────────────────────────────────────────────────

def fill_currency(universities: list):
    for uni in universities:
        uni["通貨定義"] = get_currency(uni.get("国", ""))

def fill_climate(universities: list):
    for uni in universities:
        uni["気候"] = get_climate(uni.get("国", ""))

# ────────────────────────────────────────────────────────────
# THE / QS ランキング 自動取得・キャッシュ・国内1位判定
# ────────────────────────────────────────────────────────────

# 分野別 Top30 補助データ（総合ランキングには出ない大学をカバー）
# キーワード（英語小文字・部分一致）: (THE分野最高位, QS分野最高位)
SUBJECT_RANKING_DATA = [
    ("karolinska",                          1,   1),   # 医学
    ("london business school",              None,1),   # ビジネス
    ("insead",                              None,1),   # ビジネス
    ("hec paris",                           None,1),   # ビジネス
    ("london school of hygiene",            1,   1),   # 公衆衛生
    ("university of california, los angeles",1,  5),
    ("university of california, san diego", 1,   5),
    ("university of california, san francisco",1, None),
    ("heidelberg university",               10,  None),
    ("university of amsterdam",             20,  20),
    ("university of sydney",                20,  10),
    ("university of queensland",            20,  10),
    ("monash university",                   25,  10),
    ("university of manchester",            15,  10),
    ("university of bristol",               20,  15),
    ("university of glasgow",               25,  20),
    ("seoul national university",           20,  10),
    ("korea advanced institute",            None,20),
    ("paris sciences et lettres",           15,  None),
    ("university of cape town",             25,  None),
    ("wageningen university",               1,   1),   # 農学
    ("epfl",                                1,   1),   # 工学
    ("polytechnique fédérale de lausanne",  1,   1),
    ("university of british columbia",      20,  15),
    ("kyoto university",                    None,20),
]

# ────────────────────────────────────────────────────────────
# 芸術分野ランキングデータ（2025年版・閾値50位）
#   QS  : Art & Design
#   THE : arts and humanities – art, performing arts & design
#   キーワード（英語小文字・部分一致）: (THE芸術順位, QS芸術順位)
# ────────────────────────────────────────────────────────────

ART_RANKING_DATA = [
    # ── QS Art & Design Top50 ─────────────────────────────
    ("royal college of art",                1,    1),
    ("university of the arts london",       None, 3),
    ("parsons",                             None, 4),
    ("politecnico di milano",               None, 6),
    ("pratt institute",                     None, 9),
    ("rhode island school of design",       None, 10),
    ("aalto university",                    None, 14),
    ("rmit university",                     None, 16),
    ("goldsmiths",                          None, 18),
    ("savannah college of art",             None, 19),
    ("arts university bournemouth",         None, 20),
    ("glasgow school of art",               None, 23),
    ("central saint martins",               None, 24),
    ("monash university",                   None, 25),
    ("emily carr university",               None, 22),
    ("nscad university",                    None, 28),
    ("school of the art institute",         None, 3),   # SAIC
    ("school of visual arts",               None, 30),
    ("art center college of design",        None, 35),
    ("california college of the arts",      None, 40),
    ("loughborough university",             None, 32),
    ("konstfack",                           None, 29),
    ("zurich university of the arts",       None, 45),
    ("ecal",                                None, 48),
    ("design academy eindhoven",            None, 42),
    # ── THE Arts & Humanities (art/performing arts/design) Top50 ──
    ("university of oxford",                1,    None),
    ("harvard university",                  2,    5),
    ("university of cambridge",             3,    None),
    ("stanford university",                 4,    12),
    ("university college london",           5,    7),
    ("columbia university",                 6,    None),
    ("yale university",                     7,    11),
    ("princeton university",                8,    None),
    ("new york university",                 9,    None),
    ("university of chicago",               10,   None),
    ("duke university",                     12,   None),
    ("cornell university",                  13,   13),
    ("university of edinburgh",             14,   None),
    ("king's college london",               15,   None),
    ("university of michigan",              16,   None),
    ("university of toronto",               17,   None),
    ("johns hopkins",                       18,   None),
    ("northwestern university",             19,   None),
    ("university of amsterdam",             22,   None),
    ("leiden university",                   24,   None),
    ("university of melbourne",             26,   None),
    ("university of sydney",                28,   None),
    ("australian national university",      30,   None),
    ("university of british columbia",      33,   None),
    ("mcgill university",                   36,   None),
    ("university of cape town",             40,   None),
    ("seoul national university",           42,   None),
    ("peking university",                   44,   None),
    ("tsinghua university",                 45,   None),
    ("karolinska",                          None, None),
]

# ── PDF 対象校リスト読み込み ──────────────────────────────

def load_ezo_art_list() -> set:
    """
    data_listing/ezo_art_YYYY.pdf から対象校名を読み込む。
    最新年のファイルを自動選択（ezo_art_2026.pdf → 2027.pdf → ...）。
    Returns: {大学名キーワード（小文字）} のセット
    """
    import re

    # 最新年のPDFを選択
    pdfs = sorted(DATA_LISTING_DIR.glob("ezo_art_*.pdf"), reverse=True)
    if not pdfs:
        # スクリプトと同じフォルダも探す
        pdfs = sorted(OUTPUT_DIR.glob("ezo_art_*.pdf"), reverse=True)
    if not pdfs:
        print("  [WARN] ezo_art_*.pdf が見つかりません（data_listing/ フォルダを確認してください）")
        return set()

    pdf_path = pdfs[0]
    print(f"  江副芸術対象校 PDF: {pdf_path.name}")

    # PDF テキスト抽出
    text = ""
    try:
        import pypdf
        reader = pypdf.PdfReader(str(pdf_path))
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
    except ImportError:
        try:
            import PyPDF2
            with open(pdf_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += (page.extract_text() or "") + "\n"
        except ImportError:
            print("  [WARN] PDF読み込みには pypdf が必要です: pip install pypdf")
            return set()
    except Exception as e:
        print(f"  [WARN] PDF読み込みエラー: {e}")
        return set()

    # 大学名を抽出（形式: "番号 国名 大学名 URL"）
    names = set()
    for line in text.split("\n"):
        line = line.strip()
        # 番号で始まる行: 先頭の数字と国名を除き、URLの前までを大学名とする
        m = re.match(r'^\d+\s+\S+\s+(.+?)\s+https?://\S*', line)
        if m:
            raw_name = m.group(1).strip()
            # 括弧内の補足（学部名等）を除いた短縮キーワードも登録
            short = re.sub(r'[\(（].*?[\)）]', '', raw_name).strip()
            names.add(raw_name.lower())
            if short != raw_name:
                names.add(short.lower())
        # URL なしの短い行（テキスト抽出ズレ対策）
        elif re.match(r'^\d+\s+', line):
            parts = re.split(r'\s{2,}', line, maxsplit=2)
            if len(parts) >= 3:
                names.add(parts[2].strip().lower())

    print(f"  → {len(names)} 校キーワード登録完了")
    return names

def _is_ezo_art_match(university_name: str, art_list: set) -> bool:
    """PDF対象校リストに名前が含まれるか部分一致で判定。"""
    name_lower = university_name.lower()
    # 完全一致
    if name_lower in art_list:
        return True
    # PDF側のキーワードが大学名に含まれる（または逆）
    for keyword in art_list:
        if len(keyword) >= 6 and (keyword in name_lower or name_lower in keyword):
            return True
    return False

def _parse_rank(rank_str) -> int | None:
    """'1', '=1', '201-250' などを整数に変換。範囲は下限値（最良値）を使用。"""
    if rank_str is None:
        return None
    s = str(rank_str).strip().lstrip("=").replace(",", "")
    try:
        return int(s)
    except ValueError:
        # "201-250" のような範囲 → 下限（最良値）
        if "-" in s:
            try:
                return int(s.split("-")[0])
            except ValueError:
                pass
    return None

def _name_match(name_a: str, name_b: str) -> bool:
    """2つの大学名が同一大学か判定（小文字・部分一致）。"""
    a, b = name_a.lower().strip(), name_b.lower().strip()
    return a == b or a in b or b in a

def _subject_rank(university_name: str) -> tuple[int | None, int | None]:
    """分野別補助データから (THE分野最高位, QS分野最高位) を返す。"""
    name_lower = university_name.lower()

    def _better(a, b):
        if a is None: return b
        if b is None: return a
        return min(a, b)

    the_s, qs_s = None, None
    for kw, t, q in SUBJECT_RANKING_DATA:
        if kw in name_lower:
            the_s = _better(the_s, t)
            qs_s  = _better(qs_s,  q)
    return the_s, qs_s

# ── QS ランキング取得 ──────────────────────────────────────

def fetch_qs_rankings() -> list:
    """
    QS World University Rankings を topuniversities.com API から取得。
    Returns: [{"name": str, "country": str, "rank": int}, ...]
    """
    import re
    print("    QS Rankings 取得中 (topuniversities.com)...")

    # ① ランキングページから nid を抽出
    page_url = "https://www.topuniversities.com/world-university-rankings"
    try:
        r = requests.get(page_url, headers=HEADERS_HTTP, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"    [WARN] QS ページ取得失敗: {e}")
        return []

    nid_match = (re.search(r'"nid"\s*:\s*"?(\d{5,})"?', r.text) or
                 re.search(r'nid=(\d{5,})',              r.text))
    if not nid_match:
        print("    [WARN] QS nid が見つかりませんでした")
        return []
    nid = nid_match.group(1)

    # ② API 呼び出し
    api_url = "https://www.topuniversities.com/rankings/endpoint"
    params  = {"nid": nid, "page": 0, "items_per_page": 1500,
               "tab": "indicators", "sort_by": "rank", "order_by": "asc"}
    headers = {**HEADERS_HTTP, "X-Requested-With": "XMLHttpRequest",
               "Referer": page_url}
    try:
        r = requests.get(api_url, params=params, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        print(f"    [WARN] QS API 取得失敗: {e}")
        return []

    results = []
    for item in data.get("scores", []):
        rank = _parse_rank(item.get("rank_display") or item.get("overall_rank"))
        if rank is None:
            continue
        results.append({
            "name":    item.get("title", "").strip(),
            "country": item.get("country", "").strip(),
            "rank":    rank,
        })
    print(f"    → QS {len(results)} 大学取得 (nid={nid})")
    return results

# ── THE ランキング取得 ─────────────────────────────────────

def fetch_the_rankings() -> list:
    """
    THE World University Rankings を unirank.org 経由で取得。
    Returns: [{"name": str, "country": str, "rank": int}, ...]
    """
    print("    THE Rankings 取得中 (unirank.org/the/)...")
    url = "https://www.unirank.org/the/"
    try:
        r = requests.get(url, headers=HEADERS_HTTP, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"    [WARN] THE Rankings 取得失敗: {e}")
        return []

    results = []
    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue
        rank = _parse_rank(tds[0].get_text(strip=True))
        if rank is None:
            continue
        a_tag = tds[1].find("a")
        if not a_tag:
            continue
        results.append({
            "name":    a_tag.get_text(strip=True),
            "country": tds[2].get_text(strip=True),
            "rank":    rank,
        })
    print(f"    → THE {len(results)} 大学取得")
    return results

# ── キャッシュ管理 ────────────────────────────────────────

def load_rankings_cache() -> dict:
    """
    rankings_cache.json を読み込む。
    存在しない・期限切れの場合は再取得して保存。
    """
    if RANKINGS_CACHE_FILE.exists():
        try:
            with open(RANKINGS_CACHE_FILE, encoding="utf-8") as f:
                cache = json.load(f)
            updated  = datetime.fromisoformat(cache.get("updated", "2000-01-01"))
            age_days = (datetime.now() - updated).days
            if age_days < RANKINGS_CACHE_MAX_DAYS:
                print(f"  ランキングキャッシュ使用中 (更新日: {cache['updated']} / {age_days}日前)")
                return cache
            print(f"  ランキングキャッシュが古いため再取得します ({age_days}日経過)")
        except Exception as e:
            print(f"  ランキングキャッシュ読み込みエラー ({e}) → 再取得します")

    return _refresh_rankings_cache()

def _refresh_rankings_cache() -> dict:
    """QS・THE を取得してキャッシュファイルに保存。"""
    qs_data  = fetch_qs_rankings()
    the_data = fetch_the_rankings()
    cache = {
        "updated":     datetime.now().strftime("%Y-%m-%d"),
        "qs_overall":  qs_data,
        "the_overall": the_data,
    }
    try:
        with open(RANKINGS_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        print(f"  ランキングキャッシュ保存: {RANKINGS_CACHE_FILE.name}")
    except Exception as e:
        print(f"  [WARN] キャッシュ保存失敗: {e}")
    return cache

# ── 江副判定メイン ────────────────────────────────────────

def fill_egasoe_ranking(universities: list):
    """
    江副リクルート記念財団 TRUE 条件:
      ① THE / QS 総合 Top30
      ② THE / QS 分野別（一般）Top30（補助データ）
      ③ THE / QS 芸術分野（Art & Design / art,performing arts&design）Top50
      ④ ezo_art_YYYY.pdf 指定校
      ⑤ 各国の THE 国内1位 または QS 国内1位（最大2大学/国）
    """
    from collections import defaultdict

    # ── ランキングデータ読み込み ──
    cache    = load_rankings_cache()
    qs_list  = cache.get("qs_overall",  [])
    the_list = cache.get("the_overall", [])

    # ── PDF 対象校リスト読み込み ──
    art_pdf_list = load_ezo_art_list()

    # ── 各大学に総合・分野・芸術順位をセット ──
    for uni in universities:
        name = uni.get("大学名", "")

        # QS 総合
        qs_rank = None
        for item in qs_list:
            if _name_match(name, item["name"]):
                qs_rank = item["rank"]
                break

        # THE 総合
        the_rank = None
        for item in the_list:
            if _name_match(name, item["name"]):
                the_rank = item["rank"]
                break

        # 一般分野別（Top30補助）
        the_s, qs_s = _subject_rank(name)

        # 芸術分野順位
        name_lower = name.lower()
        the_art, qs_art = None, None
        def _better(a, b):
            if a is None: return b
            if b is None: return a
            return min(a, b)
        for kw, ta, qa in ART_RANKING_DATA:
            if kw in name_lower:
                the_art = _better(the_art, ta)
                qs_art  = _better(qs_art,  qa)

        uni["THE総合"]      = the_rank
        uni["QS総合"]       = qs_rank
        uni["THE分野最高位"] = the_s
        uni["QS分野最高位"]  = qs_s
        uni["THE芸術順位"]   = the_art
        uni["QS芸術順位"]    = qs_art

    # ── 条件① ② Top30 判定 ──
    count_top30 = 0
    for uni in universities:
        if any(v is not None and v <= 30
               for v in [uni.get("THE総合"),      uni.get("QS総合"),
                         uni.get("THE分野最高位"), uni.get("QS分野最高位")]):
            uni["江副リクルート記念財団"] = "TRUE"
            count_top30 += 1

    # ── 条件③ 芸術分野 Top50 ──
    count_art50 = 0
    for uni in universities:
        if uni.get("江副リクルート記念財団") == "TRUE":
            continue
        if any(v is not None and v <= 50
               for v in [uni.get("THE芸術順位"), uni.get("QS芸術順位")]):
            uni["江副リクルート記念財団"] = "TRUE"
            count_art50 += 1
            print(f"    🎨 芸術Top50: {uni['大学名'][:40]}"
                  f"  THE芸術:{uni.get('THE芸術順位','─')}  QS芸術:{uni.get('QS芸術順位','─')}")

    # ── 条件④ PDF 指定校 ──
    count_pdf = 0
    for uni in universities:
        if uni.get("江副リクルート記念財団") == "TRUE":
            continue
        if _is_ezo_art_match(uni.get("大学名", ""), art_pdf_list):
            uni["江副リクルート記念財団"] = "TRUE"
            count_pdf += 1
            print(f"    📄 PDF指定校: {uni['大学名'][:40]}")

    # ── 条件⑤ 各国 THE/QS 国内1位 ──
    by_country = defaultdict(list)
    for uni in universities:
        by_country[uni.get("国", "")].append(uni)

    count_country1 = 0
    for country, unis in by_country.items():
        for source, key in [("THE", "THE総合"), ("QS", "QS総合")]:
            candidates = [(u[key], u) for u in unis if u.get(key) is not None]
            if not candidates:
                continue
            best_rank, best_uni = min(candidates, key=lambda x: x[0])
            if best_uni.get("江副リクルート記念財団") != "TRUE":
                best_uni["江副リクルート記念財団"] = "TRUE"
                count_country1 += 1
                print(f"    🌍 {source} 国内1位: {country} "
                      f"- {best_uni['大学名'][:35]} ({source} {best_rank}位)")

    total = sum(1 for u in universities if u.get("江副リクルート記念財団") == "TRUE")
    print(f"  → 合計 {total}大学  ("
          f"Top30: {count_top30} / 芸術Top50: {count_art50} / "
          f"PDF指定: {count_pdf} / 国内1位: {count_country1})")

# ────────────────────────────────────────────────────────────
# 奨学金対象校リスト（柳井正財団 / 笹川平和財団）
# ────────────────────────────────────────────────────────────

# 大学名として認識するキーワード（英語機関名の判別用）
_INST_KEYWORDS = (
    "university", "college", "institute", "school", "academy",
    "polytechnic", "polytechnique", "école", "ecole",   # EPFL 等の仏語機関名対応
    "minerva", "sciences", "technology",
)

def _parse_scholarship_txt(text: str) -> set:
    """
    奨学金対象校テキストから大学名キーワードセット（小文字）を抽出。
    yanai 形式 (A / B / C) と sasakawa 形式 (1行1校) の両方に対応。
    """
    names = set()
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue

        # ── yanai 形式: " / " 区切りが含まれる行 ──
        if " / " in line:
            for part in line.split(" / "):
                p = part.strip()
                if p and any(kw in p.lower() for kw in _INST_KEYWORDS):
                    names.add(p.lower())
            continue

        # ── sasakawa / YouAreWelcomeHere 形式: 英語大学名が1行に1校 ──
        lower = line.lower()
        if any(kw in lower for kw in _INST_KEYWORDS):
            # 括弧内の都市名注記 "(Garden City, New York)" 等を除去
            clean = re.sub(r'\s*[\(（][^)）]*[\)）]', '', line).strip()
            # 末尾の * を除去（YouAreWelcomeHere 形式の注記記号）
            clean = clean.rstrip('*').strip()
            names.add(clean.lower())
            if clean.lower() != lower:
                names.add(lower)   # 元表記も保持

    return names

def load_scholarship_list(prefix: str) -> set:
    """
    data_listing/{prefix}_YYYY.txt から最新年のファイルを読み込み、
    大学名キーワードセットを返す。ファイルが増えても自動で最新年を選択。
    例: yanai_2026.txt → yanai_2027.txt → ...
    """
    files = sorted(DATA_LISTING_DIR.glob(f"{prefix}_*.txt"), reverse=True)
    if not files:
        print(f"  [WARN] {prefix}_*.txt が data_listing/ に見つかりません")
        return set()

    f_path = files[0]
    print(f"  {prefix} リスト: {f_path.name}")

    try:
        text = f_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = f_path.read_text(encoding="cp932")   # Shift-JIS フォールバック
    except Exception as e:
        print(f"  [WARN] 読み込みエラー: {e}")
        return set()

    names = _parse_scholarship_txt(text)
    print(f"  → {len(names)} 校登録")
    return names

def _is_list_match(university_name: str, name_set: set) -> bool:
    """
    name_set 内のキーワードと大学名を部分一致で照合。
    キーワード長 6 文字以上のものだけを対象にして誤マッチを防ぐ。
    """
    name_lower = university_name.lower()
    if name_lower in name_set:
        return True
    for kw in name_set:
        if len(kw) < 6:
            continue
        if kw in name_lower or name_lower in kw:
            return True
    return False

def fill_scholarship_lists(universities: list):
    """
    各奨学金の対象校リストを読み込み、該当大学の各カラムを TRUE に設定する。
      - 柳井正財団
      - 笹川平和財団
      - グルーバンクロフト基金
      - YouAreWelcomeHere Scholarship
      - 東進海外進学支援制度
      - Laidlaw Scholars
      - Stamps Scholar Program
    """
    yanai_list    = load_scholarship_list("yanai")
    sasakawa_list = load_scholarship_list("sasakawa")
    grew_list     = load_scholarship_list("grew")
    yawh_list     = load_scholarship_list("YouAreWelcomeHere")
    toshin_list   = load_scholarship_list("toshin")
    laidlaw_list  = load_scholarship_list("Laidlaw")
    stamps_list   = load_scholarship_list("stamps")

    count_y = count_s = count_g = count_w = count_t = count_l = count_st = 0
    for uni in universities:
        name = uni.get("大学名", "")
        if _is_list_match(name, yanai_list):
            uni["柳井正財団"] = "TRUE"
            count_y += 1
        if _is_list_match(name, sasakawa_list):
            uni["笹川平和財団"] = "TRUE"
            count_s += 1
        if _is_list_match(name, grew_list):
            uni["グルーバンクロフト基金"] = "TRUE"
            count_g += 1
        if _is_list_match(name, yawh_list):
            uni["YouAreWelcomeHere Scholarship"] = "TRUE"
            count_w += 1
        if _is_list_match(name, toshin_list):
            uni["東進海外進学支援制度"] = "TRUE"
            count_t += 1
        if _is_list_match(name, laidlaw_list):
            uni["Laidlaw Scholars"] = "TRUE"
            count_l += 1
        if _is_list_match(name, stamps_list):
            uni["Stamps Scholar Program"] = "TRUE"
            count_st += 1

    print(f"  → 柳井正財団:{count_y}  笹川平和財団:{count_s}  グルー:{count_g}"
          f"  YouAreWelcomeHere:{count_w}  東進:{count_t}"
          f"  Laidlaw:{count_l}  Stamps:{count_st}")

# ────────────────────────────────────────────────────────────
# Excel 作成
# ────────────────────────────────────────────────────────────

# セクションごとのヘッダー背景色
SECTION_COLORS = {
    "大学名":       "1F4E79",  # 濃紺（基本情報）
    "学生数":       "1F497D",  # 紺（CDS）
    "通貨定義":     "375623",  # 深緑（費用）
    "Need総数":     "7B3F00",  # 茶（奨学金）
    "2-9":          "404040",  # 濃灰（クラスサイズ）
    "Agriculture":  "4B0082",  # 紫（学問）
    "TOEFL 総合":   "006400",  # 濃緑（語学）
    "THE総合":       "1A3A5C",  # 紺グレー（ランキング）
    "全額免除奨学金":"8B0000",  # 深赤（奨学金種別）
    "学部生のみ":   "00008B",  # 紺（大学特性）
}

def apply_input_data(universities: list):
    """
    InputDataWithPanel.py が生成した input_data.json を読み込み、
    大学名で照合して各フィールドを上書きする。

    特殊処理:
      - Need-Met == 100 の場合、"Need-Met 100%" 列を "TRUE" に設定する。
      - Need-Met != 100（または未入力）の場合は "FALSE" のまま（数値は Need-Met 列に保存）。
    """
    if not INPUT_DATA_FILE.exists():
        print(f"  ℹ️  input_data.json が見つかりません → スキップ ({INPUT_DATA_FILE})")
        return

    with open(INPUT_DATA_FILE, encoding="utf-8") as f:
        records: list[dict] = json.load(f)

    if not records:
        print("  ℹ️  input_data.json は空です → スキップ")
        return

    # 大学名 → レコード のマップ
    record_map = {r.get("大学名", "").strip(): r for r in records}

    applied = 0
    for uni in universities:
        name = uni.get("大学名", "").strip()
        rec  = record_map.get(name)
        if not rec:
            continue

        # 全フィールドを上書き（大学名・国・都市・緯度経度はスクレイピング値を優先して除外）
        SKIP_KEYS = {"大学名", "国", "都市", "緯度", "経度"}
        for key, val in rec.items():
            if key in SKIP_KEYS:
                continue
            # 種類（list型）はそのままセット
            if key == "種類":
                uni[key] = val
                continue
            # 空文字・None はスキップ（既存値を消さない）
            if val is None or val == "" or val == []:
                continue
            uni[key] = val

        # ── Need-Met 100% の自動判定 ──────────────────────────
        need_met = rec.get("Need-Met")
        try:
            need_met_num = float(need_met) if need_met not in (None, "") else None
        except (ValueError, TypeError):
            need_met_num = None

        if need_met_num is not None and need_met_num >= 100:
            uni["Need-Met 100%"] = "TRUE"
        else:
            # 明示的にFALSEは書かない（既存値があれば保持）
            if not uni.get("Need-Met 100%"):
                uni["Need-Met 100%"] = "FALSE"

        # ── Need-based Scholarship の自動判定 ─────────────────
        # 条件: Need-Met > 0 の数値あり OR Need available = TRUE
        need_available  = rec.get("Need available", False)
        need_met_filled = need_met_num is not None and need_met_num > 0

        if need_met_filled or need_available:
            uni["Need-based Scholarship"] = "TRUE"

        applied += 1

    print(f"  ✅ input_data.json から {applied} 件の大学データを反映しました")
    not_found = [r.get("大学名","") for r in records
                 if r.get("大学名","").strip() not in {u.get("大学名","").strip() for u in universities}]
    if not_found:
        print(f"  ⚠️  照合できなかった大学（スクレイピング結果に存在しない）: {len(not_found)} 件")
        for n in not_found[:10]:
            print(f"       - {n}")
        if len(not_found) > 10:
            print(f"       ... 他 {len(not_found)-10} 件")


def apply_url_backup(universities: list):
    """
    url_checker.py が生成した university_urls.json を読み込み、
    URL が空の大学に URL を補完する。
    既に URL が入っている大学は上書きしない。
    """
    if not URL_BACKUP_FILE.exists():
        print(f"  ℹ️  university_urls.json が見つかりません → スキップ ({URL_BACKUP_FILE})")
        return

    with open(URL_BACKUP_FILE, encoding="utf-8") as f:
        url_map: dict = json.load(f)

    if not url_map:
        print("  ℹ️  university_urls.json は空です → スキップ")
        return

    applied = 0
    for uni in universities:
        name = uni.get("大学名", "").strip()
        if uni.get("URL", "").strip():
            continue  # 既に URL あり → スキップ
        url = url_map.get(name, "").strip()
        if url:
            uni["URL"] = url
            applied += 1

    print(f"  ✅ university_urls.json から {applied} 件の URL を補完しました"
          f"（全 {len(url_map)} 件登録済み）")


def get_header_color(col_name: str, col_idx: int, columns: list) -> str:
    """カラムがどのセクションに属するかで色を返す。"""
    # セクション開始カラムより後ろのカラムは同じ色を継続
    current_color = "1F4E79"
    for name, color in SECTION_COLORS.items():
        try:
            section_start = columns.index(name)
        except ValueError:
            continue
        if col_idx - 1 >= section_start:
            current_color = color
    return current_color


def build_excel(universities: list, output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "大学データ"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── ヘッダー行
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        color = get_header_color(col_name, col_idx, COLUMNS)
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = PatternFill("solid", start_color=color, end_color=color)
        cell.alignment = header_align

    ws.row_dimensions[1].height = 45

    data_font = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
    cb_fill   = PatternFill("solid", start_color="F0F8E8", end_color="F0F8E8")  # チェックボックス列の薄緑

    # チェックボックス列インデックス（1-based）
    cb_col_indices = {
        col_idx for col_idx, col_name in enumerate(COLUMNS, start=1)
        if col_name in CHECKBOX_COLUMNS
    }

    # ── データ行
    max_row = len(universities) + 1
    for row_idx, uni in enumerate(universities, start=2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = uni.get(col_name, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font

            if col_idx in cb_col_indices:
                # チェックボックス列: 薄緑背景
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = cb_fill
                if value is None:
                    cell.value = "FALSE"
            else:
                cell.alignment = Alignment(
                    horizontal="left" if col_name in ("大学名", "URL") else "center",
                    vertical="center",
                )
                if is_alt:
                    cell.fill = alt_fill

    # DataValidation: Python 3.14 + openpyxl の add_data_validation() フリーズバグを回避するため
    # 内部リストへ直接 append する（sqref の MultiCellRange.__contains__ の無限ループを回避）
    for cb_col_idx in cb_col_indices:
        letter = get_column_letter(cb_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
            sqref=f"{letter}2:{letter}{max_row}",
        )
        dv.showDropDown = False
        ws.data_validations.dataValidation.append(dv)

    # ── 列幅
    col_widths = {
        "大学名": 42, "国": 16, "都市": 18, "URL": 52,
        "緯度": 13, "経度": 13, "気候": 14, "通貨定義": 12,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 13)

    # ── フリーズ: 1行目 + A列
    ws.freeze_panes = "B2"

    # ── オートフィルター
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(output_path)
    print(f"  ✅ Excel: {output_path.name}  ({len(universities)}大学 / {len(COLUMNS)}カラム)")


# ────────────────────────────────────────────────────────────
# CSV 作成
# ────────────────────────────────────────────────────────────

def build_csv(universities: list, output_path: Path):
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for uni in universities:
            row = {}
            for col in COLUMNS:
                val = uni.get(col, "")
                # チェックボックス列はデフォルト FALSE
                if col in CHECKBOX_COLUMNS and (val is None or val == ""):
                    val = "FALSE"
                row[col] = val
            writer.writerow(row)
    print(f"  ✅ CSV:   {output_path.name}")


# ────────────────────────────────────────────────────────────
# メイン
# ────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  大学データセット 自動生成スクリプト v2")
    print("=" * 55)

    if TEST_MODE:
        print("\n[テストモード] サンプルデータを使用します")
        all_universities = get_sample_data()
    else:
        all_universities = []

        # ── STEP 1: スクレイピング（国ごとに保存して中断再開対応）
        print(f"\n[STEP 1] unirank.org スクレイピング ({len(COUNTRIES)}カ国)")
        scraped: dict = {}
        if SCRAPE_PROGRESS_FILE.exists():
            with open(SCRAPE_PROGRESS_FILE, encoding="utf-8") as f:
                scraped = json.load(f)
            print(f"  → スクレイピング進捗を読み込みました ({len(scraped)}カ国済み)")

        for country_name, url in COUNTRIES.items():
            if country_name in scraped:
                unis = scraped[country_name]
                all_universities.extend(unis)
                print(f"  {country_name}: {len(unis)}大学 [スキップ（保存済み）]")
                continue
            print(f"  {country_name}: {url}")
            unis = scrape_unirank(country_name, url)
            all_universities.extend(unis)
            # 即座に保存（中断対策）
            scraped[country_name] = unis
            with open(SCRAPE_PROGRESS_FILE, "w", encoding="utf-8") as f:
                json.dump(scraped, f, ensure_ascii=False)
            time.sleep(2)

        print(f"\n  合計 {len(all_universities)} 大学を取得")

        if not all_universities:
            print("\n[警告] データが取得できませんでした。TEST_MODE = True で動作確認してください。")
            return

        # ── STEP 1.5: 同一国内の重複大学名に連番付加（ジオコーディング前に実施）
        print(f"\n[STEP 1.5] 重複大学名チェック...")
        deduplicate_names(all_universities)

        # ── STEP 2: ジオコーディング（大学名+国名をキーにして中断再開対応）
        print(f"\n[STEP 2] ジオコーディング (Nominatim / OpenStreetMap)")
        print(f"  進捗ファイル: {PROGRESS_FILE}")

        progress = {}
        if PROGRESS_FILE.exists():
            with open(PROGRESS_FILE, encoding="utf-8") as f:
                progress = json.load(f)
            done = sum(1 for v in progress.values() if v["lat"] is not None)
            print(f"  → 前回の進捗を読み込みました ({len(progress)}件保存済み、うち成功{done}件)")

        for i, uni in enumerate(all_universities, start=1):
            # キーは「国名||大学名」で一意性を確保（同名別国の混同を防ぐ）
            key = f"{uni['国']}||{uni['大学名']}"
            if key in progress:
                uni["緯度"] = progress[key]["lat"]
                uni["経度"] = progress[key]["lon"]
                print(f"  ({i:5d}/{len(all_universities)}) {uni['大学名'][:40]}... [スキップ]")
                continue

            print(f"  ({i:5d}/{len(all_universities)}) {uni['大学名'][:40]}...", end=" ", flush=True)
            lat, lon = geocode(uni["大学名"], uni["都市"], uni["国"])
            uni["緯度"] = lat
            uni["経度"] = lon

            # 即座に進捗を保存（中断対策）
            progress[key] = {"lat": lat, "lon": lon}
            with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
                json.dump(progress, f, ensure_ascii=False)

            print(f"lat={lat:.4f}, lon={lon:.4f}" if lat else "取得失敗")

        print(f"\n  ジオコーディング完了。進捗ファイルは {PROGRESS_FILE.name} に保存されています。")
        print(f"  全件完了後に不要であれば削除してください。")

    # テストモード時のみ重複チェック（通常モードはSTEP 1.5で実施済み）
    if TEST_MODE:
        print("\n[STEP 1.5] 重複大学名チェック...")
        deduplicate_names(all_universities)

    # 気候自動補完
    print("\n[STEP 3a] 気候を自動補完...")
    fill_climate(all_universities)
    for uni in all_universities:
        print(f"  {uni['国']:12s}  →  気候: {uni['気候'] or '(未定義)'}")

    # 江副リクルート記念財団：THE/QS Top30 自動判定
    print("\n[STEP 3b] 江副リクルート記念財団（THE/QS Top30）を自動判定...")
    fill_egasoe_ranking(all_universities)
    for uni in all_universities:
        if uni.get("江副リクルート記念財団") == "TRUE":
            print(f"  ✅ {uni['大学名'][:40]:40s}"
                  f"  THE総合:{str(uni.get('THE総合','─')):>4s}"
                  f"  QS総合:{str(uni.get('QS総合','─')):>4s}"
                  f"  THE分野:{str(uni.get('THE分野最高位','─')):>4s}"
                  f"  QS分野:{str(uni.get('QS分野最高位','─')):>4s}")

    # 柳井正財団・笹川平和財団 対象校リスト判定
    print("\n[STEP 3c] 柳井正財団・笹川平和財団 対象校を自動判定...")
    fill_scholarship_lists(all_universities)

    # 通貨自動補完
    print("\n[STEP 3d] 通貨定義を自動補完...")
    fill_currency(all_universities)
    for uni in all_universities:
        print(f"  {uni['大学名'][:30]:30s}  →  {uni['通貨定義']}")

    # 手動入力データ（InputDataWithPanel.py）を反映
    print("\n[STEP 3e] 手動入力データ（input_data.json）を反映...")
    apply_input_data(all_universities)

    # URL バックアップから補完（url_checker.py で生成）
    print("\n[STEP 3f] URL バックアップ（university_urls.json）を反映...")
    apply_url_backup(all_universities)

    # ファイル出力
    stem = get_output_stem()
    xlsx_path = stem.with_suffix(".xlsx")
    csv_path  = stem.with_suffix(".csv")

    print(f"\n[STEP 4] ファイル出力")
    build_excel(all_universities, xlsx_path)
    build_csv(all_universities, csv_path)

    # JSON バックアップ
    json_path = stem.with_suffix(".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_universities, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON:  {json_path.name}  (バックアップ)")

    print(f"\n完了！ 出力先: {OUTPUT_DIR.resolve()}")
    print(f"  ※ 全件完了したら scrape_progress.json と geocode_progress.json は削除して構いません")


if __name__ == "__main__":
    main()