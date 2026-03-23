"""
Excel → CSV 変換ツール
- ファイル選択ダイアログ付き
- 色・チェックボックス情報をメタデータCSVとして別途出力
- 出力先はこのスクリプトと同じフォルダー
"""

import os
import sys
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl が見つかりません。インストールしてください: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ---- Excel解析ユーティリティ ----

def rgb_from_argb(argb: str) -> str:
    """ARGBカラー文字列 (例: 'FFAABBCC') を '#AABBCC' に変換"""
    if argb and len(argb) == 8:
        return f"#{argb[2:]}"
    if argb and len(argb) == 6:
        return f"#{argb}"
    return ""

def get_cell_bg_color(cell) -> str:
    fill = cell.fill
    if fill and fill.fill_type not in (None, "none"):
        fg = fill.fgColor
        if fg:
            if fg.type == "rgb" and fg.rgb not in ("00000000", "FFFFFFFF", ""):
                return rgb_from_argb(fg.rgb)
            if fg.type == "theme":
                return f"theme:{fg.theme}"
    return ""

def get_cell_font_color(cell) -> str:
    font = cell.font
    if font and font.color:
        c = font.color
        if c.type == "rgb" and c.rgb not in ("00000000", "FF000000", ""):
            return rgb_from_argb(c.rgb)
        if c.type == "theme":
            return f"theme:{c.theme}"
    return ""

def is_checkbox_like(value) -> bool:
    """チェックボックス的な値を検出（TRUE/FALSE, ✓, ☑ など）"""
    if isinstance(value, bool):
        return True
    if isinstance(value, str):
        return value.strip() in ("TRUE", "FALSE", "✓", "✗", "☑", "☐", "×", "〇", "●", "○")
    return False


def convert_excel_to_csv(xlsx_path: str, progress_callback=None):
    """
    Excelファイルを変換してCSVとメタデータCSVを生成する。
    Returns: list of (output_csv_path, metadata_csv_path) per sheet
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    results = []

    total_sheets = len(wb.sheetnames)
    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]

        # シートが複数あればサフィックスを付ける
        if total_sheets == 1:
            suffix = ""
        else:
            safe_sheet = sheet_name.replace("/", "_").replace("\\", "_")
            suffix = f"_{safe_sheet}"

        csv_path = os.path.join(SCRIPT_DIR, f"{base_name}{suffix}.csv")
        meta_path = os.path.join(SCRIPT_DIR, f"{base_name}{suffix}_metadata.csv")

        data_rows = []
        meta_rows = []  # (行, 列, 背景色, 文字色, チェックボックス, 元の値)

        for row in ws.iter_rows():
            row_values = []
            for cell in row:
                val = cell.value

                # 数式セルは data_only=True で計算済み値が入る
                row_values.append("" if val is None else str(val))

                # メタデータ収集
                bg = get_cell_bg_color(cell)
                fc = get_cell_font_color(cell)
                ck = is_checkbox_like(val)
                if bg or fc or ck:
                    col_letter = get_column_letter(cell.column)
                    meta_rows.append({
                        "セル": f"{col_letter}{cell.row}",
                        "行": cell.row,
                        "列": col_letter,
                        "背景色": bg,
                        "文字色": fc,
                        "チェックボックス的な値": "はい" if ck else "",
                        "元の値": "" if val is None else str(val),
                    })

            data_rows.append(row_values)

        # --- CSVを書き出し ---
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(data_rows)

        # --- メタデータCSVを書き出し ---
        if meta_rows:
            meta_fields = ["セル", "行", "列", "背景色", "文字色", "チェックボックス的な値", "元の値"]
            with open(meta_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=meta_fields)
                writer.writeheader()
                writer.writerows(meta_rows)
        else:
            meta_path = None  # メタデータなし

        results.append((csv_path, meta_path, sheet_name))

        if progress_callback:
            progress_callback(sheet_idx + 1, total_sheets)

    wb.close()
    return results


# ---- GUI ----

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel → CSV 変換ツール")
        self.resizable(False, False)
        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 12, "pady": 6}

        # ファイル選択
        frame_file = tk.LabelFrame(self, text="Excelファイルを選択", **pad)
        frame_file.pack(fill="x", **pad)

        self.file_var = tk.StringVar()
        entry = tk.Entry(frame_file, textvariable=self.file_var, width=52, state="readonly")
        entry.pack(side="left", padx=(4, 4), pady=4)

        btn_browse = tk.Button(frame_file, text="参照…", command=self._browse)
        btn_browse.pack(side="left", pady=4)

        # 出力先表示
        frame_out = tk.LabelFrame(self, text="出力先フォルダー", **pad)
        frame_out.pack(fill="x", **pad)
        tk.Label(frame_out, text=SCRIPT_DIR, anchor="w", foreground="#444").pack(
            fill="x", padx=4, pady=4
        )

        # 変換ボタン
        self.btn_convert = tk.Button(
            self, text="CSVに変換", command=self._convert,
            bg="#2563eb", fg="white", font=("", 11, "bold"),
            padx=16, pady=6, state="disabled"
        )
        self.btn_convert.pack(**pad)

        # プログレスバー
        self.progress = ttk.Progressbar(self, mode="determinate", length=400)
        self.progress.pack(**pad)

        # ログ
        frame_log = tk.LabelFrame(self, text="ログ", **pad)
        frame_log.pack(fill="both", expand=True, **pad)
        self.log = tk.Text(frame_log, height=10, state="disabled", wrap="word",
                           font=("Courier", 9))
        self.log.pack(fill="both", expand=True, padx=4, pady=4)

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[("Excelファイル", "*.xlsx *.xlsm *.xltx *.xltm"), ("すべて", "*.*")]
        )
        if path:
            self.file_var.set(path)
            self.btn_convert.config(state="normal")
            self._log_clear()

    def _log(self, msg: str):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.config(state="disabled")
        self.update_idletasks()

    def _log_clear(self):
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")

    def _convert(self):
        xlsx_path = self.file_var.get()
        if not xlsx_path or not os.path.isfile(xlsx_path):
            messagebox.showerror("エラー", "有効なExcelファイルを選択してください。")
            return

        self.btn_convert.config(state="disabled")
        self.progress["value"] = 0
        self._log_clear()
        self._log(f"変換開始: {os.path.basename(xlsx_path)}")

        try:
            def on_progress(done, total):
                self.progress["maximum"] = total
                self.progress["value"] = done

            results = convert_excel_to_csv(xlsx_path, progress_callback=on_progress)

            for csv_path, meta_path, sheet_name in results:
                self._log(f"\n[シート: {sheet_name}]")
                self._log(f"  ✔ CSV出力      → {os.path.basename(csv_path)}")
                if meta_path:
                    self._log(f"  ✔ メタデータ   → {os.path.basename(meta_path)}")
                    self._log(f"    （色・チェックボックス情報を記録）")
                else:
                    self._log(f"  ℹ メタデータなし（色・チェックボックスは検出されませんでした）")

            self._log("\n✅ 変換完了！")
            messagebox.showinfo("完了", "変換が完了しました。\n出力先を確認してください。")

        except Exception as e:
            self._log(f"\n❌ エラー: {e}")
            messagebox.showerror("変換エラー", str(e))

        finally:
            self.btn_convert.config(state="normal")


if __name__ == "__main__":
    app = App()
    app.mainloop()