"""
listFromWiki.py
Wikipedia から日本の47都道府県の高校一覧を取得して Excel に出力する。
変更（新規・廃校）を検出して Discord に通知する。

出力先: GetData/Highschool/listData/YYYY_MM_DD.xlsx
         同日に複数回実行した場合は YYYY_MM_DD(1).xlsx, (2).xlsx ...
"""

import re
import sys
import time
import json
import os
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import URLError

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# ── 設定 ──────────────────────────────────────────────────
load_dotenv(Path(__file__).parent.parent.parent / ".env")

WIKI_BASE     = "https://ja.wikipedia.org"
MAIN_PAGE_URL = WIKI_BASE + "/wiki/%E6%97%A5%E6%9C%AC%E3%81%AE%E9%AB%98%E7%AD%89%E5%AD%A6%E6%A0%A1%E4%B8%80%E8%A6%A7"
OUTPUT_DIR    = Path(__file__).parent / "listData"
SNAPSHOT_FILE   = Path(__file__).parent / "last_snapshot.json"
PREF_URLS_FILE  = Path(__file__).parent / "pref_urls.json"
MANUAL_DATA_FILE = Path(__file__).parent / "manual_data.json"

DISCORD_WEBHOOK = os.getenv("DISCORD_HIGHSCHOOL_NAME_CHANGE_WEBHOOK")

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; WillAbroadBot/1.0)"}
SLEEP_SEC = 1.2  # リクエスト間隔（Wikipedia に優しく）

# 都道府県の正しい読み名マッピング（URL デコードから取得できない場合のフォールバック）
PREF_PATTERN = re.compile(
    r"(北海道|青森|岩手|宮城|秋田|山形|福島|茨城|栃木|群馬|埼玉|千葉|東京|神奈川"
    r"|新潟|富山|石川|福井|山梨|長野|岐阜|静岡|愛知|三重|滋賀|京都|大阪|兵庫"
    r"|奈良|和歌山|鳥取|島根|岡山|広島|山口|徳島|香川|愛媛|高知|福岡|佐賀|長崎"
    r"|熊本|大分|宮崎|鹿児島|沖縄)(都|道|府|県)?"
)

# ── ファイル名生成 ─────────────────────────────────────────
def make_output_path() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    base = datetime.now().strftime("%Y_%m_%d")
    path = OUTPUT_DIR / f"{base}.xlsx"
    if not path.exists():
        return path
    n = 1
    while True:
        path = OUTPUT_DIR / f"{base}({n}).xlsx"
        if not path.exists():
            return path
        n += 1

# ── HTTP 取得 ─────────────────────────────────────────────
def fetch_html(url: str, retry: int = 3) -> str:
    for attempt in range(retry):
        try:
            req = Request(url, headers=HEADERS)
            return urlopen(req, timeout=20).read().decode("utf-8")
        except URLError as e:
            if attempt < retry - 1:
                time.sleep(3)
            else:
                raise e

# ── 都道府県リンク抽出 ────────────────────────────────────
def get_prefecture_links() -> list[tuple[str, str]]:
    """メインページから各都道府県の高校一覧ページのリンクを取得する。
    Returns: [(都道府県名, URL), ...]
    """
    import urllib.parse
    html = fetch_html(MAIN_PAGE_URL)

    # Wikipedia は href を %XX エンコードしたまま出力するので raw regex で抽出
    raw_hrefs = re.findall(
        r'href="(/wiki/[^"]*%E9%AB%98%E7%AD%89%E5%AD%A6%E6%A0%A1%E4%B8%80%E8%A6%A7)"',
        html
    )
    raw_hrefs = list(dict.fromkeys(raw_hrefs))

    results = []
    seen_prefs = set()
    skip_keywords = ["日本の高等学校", "特別:", "ノート:", "利用者:", "カテゴリ:"]

    for href in raw_hrefs:
        decoded = urllib.parse.unquote(href)  # 例: /wiki/北海道高等学校一覧
        if any(k in decoded for k in skip_keywords):
            continue
        m = PREF_PATTERN.search(decoded)
        if not m:
            continue
        pref = m.group(0)
        if pref in seen_prefs:
            continue
        results.append((pref, WIKI_BASE + href))
        seen_prefs.add(pref)

    return results

# ── 高校一覧抽出 ──────────────────────────────────────────
def get_schools_from_page(pref: str, url: str) -> list[dict]:
    """都道府県別高校一覧ページから高校名・URL を取得する。"""
    html = fetch_html(url)
    soup = BeautifulSoup(html, "html.parser")
    # mw-content-text → 最初の div 子要素が信頼できるコンテナ
    mw_text = soup.find("div", id="mw-content-text")
    if mw_text:
        first_div = mw_text.find("div")
        content = first_div if first_div else mw_text
    else:
        content = soup.find("div", {"class": "mw-parser-output"}) or soup

    schools = []
    seen_names = set()

    for li in content.find_all("li"):
        # 高校らしいリンクが含まれているかチェック
        a = li.find("a", href=True)
        if not a:
            # リンクなし（廃校・記事なし）はテキストから名前を取得
            text = li.get_text(strip=True)
            text = re.sub(r"[（(].*?[）)]", "", text).strip()
            text = re.sub(r"\s+", " ", text).strip()
            if _is_valid_school_name(text) and text not in seen_names:
                schools.append({"name": text, "pref": pref, "url": ""})
                seen_names.add(text)
            continue

        href = a["href"]
        # /wiki/ または //ja.wikipedia.org/wiki/ どちらも対応
        wiki_path = None
        if href.startswith("/wiki/"):
            wiki_path = href
        elif href.startswith("//ja.wikipedia.org/wiki/"):
            wiki_path = href[len("//ja.wikipedia.org"):]  # → /wiki/...
        else:
            continue  # 外部リンク等は除外

        if any(x in wiki_path for x in ["特別:", "ノート:", "利用者:", "カテゴリ:", "Wikipedia:", "Help:", "Portal:"]):
            continue

        name = a.get_text(strip=True)
        name = re.sub(r"[（(].*?[）)]", "", name).strip()
        name = re.sub(r"\s+", " ", name).strip()

        if not _is_valid_school_name(name):
            continue
        if name in seen_names:
            continue

        school_url = WIKI_BASE + wiki_path
        schools.append({"name": name, "pref": pref, "url": school_url})
        seen_names.add(name)

    return schools

def _is_valid_school_name(name: str) -> bool:
    """高校名として有効かチェック。"""
    if len(name) < 3 or len(name) > 40:
        return False
    # 文章（句点・読点を含む）は除外
    if re.search(r"[。、…]", name):
        return False
    # ナビゲーション系の文字列を除外
    skip = ["一覧", "高等学校一覧", "Wikipedia", "曖昧さ回避", "この項目",
            "出典", "外部リンク", "関連項目", "脚注", "注釈", "参考文献",
            "廃止", "Category", "Portal", "編集", "表示", "→", "←"]
    if any(s in name for s in skip):
        return False
    # 学校名は「高等学校」「高校」「中等教育学校」で終わるか含む
    if not re.search(r"(高等学校|高校|中等教育学校)", name):
        return False
    return True

# ── Discord 通知 ──────────────────────────────────────────
def send_discord_change(changes: list[dict]) -> None:
    if not DISCORD_WEBHOOK:
        print("[WARN] DISCORD_HIGHSCHOOL_NAME_CHANGE_WEBHOOK が未設定のため通知をスキップ")
        return
    if not changes:
        return

    new_schools = [c for c in changes if c["type"] == "new"]
    closed      = [c for c in changes if c["type"] == "closed"]

    embeds = []

    if new_schools:
        desc = "\n".join(f"• {c['name']}（{c['pref']}）" for c in new_schools[:20])
        if len(new_schools) > 20:
            desc += f"\n… 他 {len(new_schools)-20} 件"
        embeds.append({
            "title": f"🏫 新規高校 {len(new_schools)} 件",
            "description": desc,
            "color": 0x3dba6e,
        })

    if closed:
        desc = "\n".join(f"• {c['name']}（{c['pref']}）" for c in closed[:20])
        if len(closed) > 20:
            desc += f"\n… 他 {len(closed)-20} 件"
        embeds.append({
            "title": f"🚪 廃校・削除 {len(closed)} 件",
            "description": desc,
            "color": 0xe04444,
        })

    payload = {
        "username": "Will Abroad — 高校変更通知",
        "embeds": embeds,
    }
    try:
        r = requests.post(DISCORD_WEBHOOK, json=payload, timeout=10)
        r.raise_for_status()
        print(f"[Discord] 変更通知を送信しました ({len(changes)} 件)")
    except Exception as e:
        print(f"[WARN] Discord 通知失敗: {e}")

# ── スナップショット比較 ──────────────────────────────────
def load_pref_urls() -> list[tuple[str, str]]:
    """キャッシュされた都道府県URLを読み込む。"""
    if not PREF_URLS_FILE.exists():
        return []
    try:
        data = json.loads(PREF_URLS_FILE.read_text(encoding="utf-8"))
        return [(d["pref"], d["url"]) for d in data]
    except Exception:
        return []

def save_pref_urls(links: list[tuple[str, str]]) -> None:
    data = [{"pref": p, "url": u} for p, u in links]
    PREF_URLS_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def load_snapshot() -> set[str]:
    if not SNAPSHOT_FILE.exists():
        return set()
    try:
        data = json.loads(SNAPSHOT_FILE.read_text(encoding="utf-8"))
        return set(data)
    except Exception:
        return set()

def save_snapshot(names: set[str]) -> None:
    SNAPSHOT_FILE.write_text(
        json.dumps(sorted(names), ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def detect_changes(current: list[dict], prev_names: set[str]) -> list[dict]:
    current_names = {s["name"] for s in current}
    new_schools    = [{"type": "new",    "name": s["name"], "pref": s["pref"]}
                      for s in current if s["name"] not in prev_names]
    closed_schools = [{"type": "closed", "name": n,        "pref": "—"}
                      for n in prev_names if n not in current_names]
    return new_schools + closed_schools

# ── 手動データ 読み込み / 保存 ──────────────────────────────
def load_manual_data() -> dict:
    """manual_data.json を読み込む。
    形式: { "高校名": {"url": "...", "confirmed": True/False, "overseas": True/False} }
    """
    if not MANUAL_DATA_FILE.exists():
        return {}
    try:
        return json.loads(MANUAL_DATA_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

def extract_manual_from_excel(excel_path: Path) -> dict:
    """既存の Excel ファイルから手動入力列（C・D・E）を読み取る。"""
    from openpyxl import load_workbook
    manual = {}
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name      = str(row[0]).strip()
            url       = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            confirmed = row[3] if len(row) > 3 else False
            overseas  = row[4] if len(row) > 4 else False
            # TRUE/FALSE 文字列 → bool
            if isinstance(confirmed, str):
                confirmed = confirmed.upper() == "TRUE"
            if isinstance(overseas, str):
                overseas = overseas.upper() == "TRUE"
            if url or confirmed or overseas:
                manual[name] = {
                    "url":       url,
                    "confirmed": bool(confirmed),
                    "overseas":  bool(overseas),
                }
        wb.close()
    except Exception as e:
        print(f"[WARN] 既存Excelの読み取りスキップ: {e}")
    return manual

def merge_and_save_manual(new_entries: dict) -> dict:
    """既存の manual_data.json と new_entries をマージして保存する。
    新しい値が空でない場合のみ上書き。
    """
    existing = load_manual_data()
    for name, data in new_entries.items():
        if name not in existing:
            existing[name] = data
        else:
            # URL: 新しい値があれば上書き
            if data.get("url"):
                existing[name]["url"] = data["url"]
            # チェック: True になったら上書き（False への上書きはしない）
            if data.get("confirmed"):
                existing[name]["confirmed"] = True
            if data.get("overseas"):
                existing[name]["overseas"] = True
    MANUAL_DATA_FILE.write_text(
        json.dumps(existing, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return existing

# ── Excel 出力 ────────────────────────────────────────────
def save_excel(schools: list[dict], output_path: Path, manual: dict) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation
    from collections import Counter

    year = datetime.now().year
    wb = Workbook()
    ws = wb.active
    ws.title = "高校一覧"

    # ── ヘッダー ──
    header_fill = PatternFill(fill_type="solid", fgColor="2A2D3E")
    header_font = Font(bold=True, color="E8E8F0", size=11)
    headers    = ["高校名", "都道府県", "公式サイトURL", f"{year}年度確認済", "海外大輩出"]
    col_widths = [40,        14,          52,              16,               12]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20

    # ── チェックボックス用データバリデーション（D・E 列）──
    dv_check = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False, showDropDown=False)
    dv_check.sqref = f"D2:E{len(schools) + 1}"
    ws.add_data_validation(dv_check)

    # ── データ行 ──
    white_fill     = PatternFill(fill_type="solid", fgColor="FFFFFF")   # 奇数行：白
    yellow_fill    = PatternFill(fill_type="solid", fgColor="FFFDE7")   # 偶数行：薄い黄色
    confirmed_fill = PatternFill(fill_type="solid", fgColor="C8F0D8")   # 確認済：薄緑
    overseas_fill  = PatternFill(fill_type="solid", fgColor="C8E4F8")   # 海外大輩出：薄青
    black_font     = Font(color="000000")
    true_font      = Font(bold=True, color="1A7A3C")
    url_font       = Font(color="1155CC", underline="single")

    for row_idx, s in enumerate(schools, start=2):
        name = s["name"]
        m    = manual.get(name, {})

        url       = m.get("url", "")
        confirmed = m.get("confirmed", False)
        overseas  = m.get("overseas",  False)

        # 行の背景（優先順: 確認済 > 海外大輩出 > 偶数行黄 > 奇数行白）
        if confirmed:
            row_fill = confirmed_fill
        elif overseas:
            row_fill = overseas_fill
        elif row_idx % 2 == 0:
            row_fill = yellow_fill
        else:
            row_fill = white_fill

        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = row_fill
            ws.cell(row=row_idx, column=col).font = black_font

        # A: 高校名
        ws.cell(row=row_idx, column=1, value=name)
        # B: 都道府県
        ws.cell(row=row_idx, column=2, value=s["pref"])
        # C: 公式URL
        c_cell = ws.cell(row=row_idx, column=3, value=url)
        if url:
            c_cell.hyperlink = url
            c_cell.font = url_font
        # D: 確認済
        d_cell = ws.cell(row=row_idx, column=4, value="TRUE" if confirmed else "FALSE")
        if confirmed:
            d_cell.font = true_font
        d_cell.alignment = Alignment(horizontal="center")
        # E: 海外大輩出
        e_cell = ws.cell(row=row_idx, column=5, value="TRUE" if overseas else "FALSE")
        if overseas:
            e_cell.font = true_font
        e_cell.alignment = Alignment(horizontal="center")

    # ウィンドウ枠固定
    ws.freeze_panes = "A2"

    # ── 統計シート ──
    ws2 = wb.create_sheet("都道府県別")
    ws2.cell(row=1, column=1, value="都道府県").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="高校数").font   = Font(bold=True)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 10
    pref_counts = Counter(s["pref"] for s in schools)
    for row_idx, (pref, count) in enumerate(sorted(pref_counts.items()), start=2):
        ws2.cell(row=row_idx, column=1, value=pref)
        ws2.cell(row=row_idx, column=2, value=count)

    wb.save(output_path)
    confirmed_count = sum(1 for s in schools if manual.get(s["name"], {}).get("confirmed"))
    overseas_count  = sum(1 for s in schools if manual.get(s["name"], {}).get("overseas"))
    url_count       = sum(1 for s in schools if manual.get(s["name"], {}).get("url"))
    print(f"[保存] {output_path}")
    print(f"       {len(schools):,} 校  |  URL入力済: {url_count}  |  確認済: {confirmed_count}  |  海外大輩出: {overseas_count}")

# ── メイン ────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("日本の高校一覧 — Wikipedia スクレイピング")
    print(f"開始: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 都道府県リンク取得（キャッシュ優先）
    print("\n[1] 都道府県リンクを取得中...")
    cached = load_pref_urls()
    if cached and len(cached) >= 47:
        pref_links = cached
        print(f"    → キャッシュから {len(pref_links)} 都道府県を読み込み（pref_urls.json）")
    else:
        pref_links = get_prefecture_links()
        save_pref_urls(pref_links)
        print(f"    → Wikipedia から {len(pref_links)} 都道府県を取得・キャッシュ保存")
    if len(pref_links) < 40:
        print("[WARN] 47都道府県に満たない場合があります。手動確認を推奨。")

    # 各都道府県の高校一覧取得
    print("\n[2] 各都道府県の高校一覧を取得中...")
    all_schools: list[dict] = []
    for i, (pref, url) in enumerate(pref_links, start=1):
        print(f"    ({i:2d}/{len(pref_links)}) {pref} ...", end=" ", flush=True)
        try:
            schools = get_schools_from_page(pref, url)
            all_schools.extend(schools)
            print(f"{len(schools)} 校")
        except Exception as e:
            print(f"エラー: {e}")
        time.sleep(SLEEP_SEC)

    print(f"\n    合計: {len(all_schools):,} 校")

    # 変更検出
    print("\n[3] 前回スナップショットと比較中...")
    prev_names = load_snapshot()
    if prev_names:
        changes = detect_changes(all_schools, prev_names)
        new_count    = sum(1 for c in changes if c["type"] == "new")
        closed_count = sum(1 for c in changes if c["type"] == "closed")
        print(f"    新規: {new_count} 校  廃校・削除: {closed_count} 校")
        if changes:
            send_discord_change(changes)
        else:
            print("    変更なし")
    else:
        print("    初回実行のためスナップショットを作成します（通知なし）")

    # スナップショット保存
    current_names = {s["name"] for s in all_schools}
    save_snapshot(current_names)

    # 手動データ復元
    print("\n[4] 手動入力データを読み込み中...")
    # まず既存の Excel から C・D・E 列を抽出
    existing_excels = sorted(OUTPUT_DIR.glob("*.xlsx"))
    excel_manual: dict = {}
    if existing_excels:
        latest_excel = existing_excels[-1]
        print(f"    既存Excel: {latest_excel.name} から手動データを抽出")
        excel_manual = extract_manual_from_excel(latest_excel)
        print(f"    → {len(excel_manual)} 件のデータを検出")
    # JSON バックアップとマージ（Excel > JSON で上書き）
    manual = merge_and_save_manual(excel_manual)
    filled = sum(1 for v in manual.values() if v.get("url") or v.get("confirmed") or v.get("overseas"))
    print(f"    manual_data.json に {filled} 件の手動データを保持")

    # Excel 出力
    print("\n[5] Excel ファイルを出力中...")
    output_path = make_output_path()
    save_excel(all_schools, output_path, manual)

    print("\n完了！")
    print(f"出力ファイル: {output_path}")
    print(f"合計校数: {len(all_schools):,} 校")

if __name__ == "__main__":
    main()
