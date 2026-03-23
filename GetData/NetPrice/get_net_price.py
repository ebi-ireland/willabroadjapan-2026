import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
import time

# ==============================================================
# ⚠️  College Navigator URL設定
# URLが変更された場合、以下のURLを更新してください。
COLLEGE_NAVIGATOR_URL = "https://nces.ed.gov/collegenavigator/"
# ==============================================================

def px_to_col_width(px):
    return px / 7.5

def get_unique_filename(directory, base_name):
    filepath = os.path.join(directory, f"{base_name}.xlsx")
    if not os.path.exists(filepath):
        return filepath
    counter = 1
    while True:
        filepath = os.path.join(directory, f"{base_name}({counter}).xlsx")
        if not os.path.exists(filepath):
            return filepath
        counter += 1

def create_workbook(college_data, filepath):
    """
    Sheet1「年度最新版」: 大学名・種類・最新年度の平均＋収入帯別データ
    Sheet2「大学別年度変移」: 1行目=大学名、A列=ラベル、全年度データを列方向に展開
    """
    wb = Workbook()

    # --- Sheet1: 年度最新版 ---
    ws1 = wb.active
    ws1.title = "年度最新版"

    headers1 = [
        "大学名",           # A
        "種類",             # B
        "平均",             # C
        "$0-$30,000",       # D
        "$30,001-$48,000",  # E
        "$48,001-$75,000",  # F
        "$75,001-$110,000", # G
        "$110,001+",        # H
    ]
    center = Alignment(horizontal="center")
    for col, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.alignment = center

    col_widths1 = {"A": 240, "B": 90, "C": 150, "D": 150, "E": 150, "F": 150, "G": 150, "H": 150}
    for col_letter, px in col_widths1.items():
        ws1.column_dimensions[col_letter].width = px_to_col_width(px)

    for row_idx, (col_a, col_b) in enumerate(college_data, start=2):
        ws1.cell(row=row_idx, column=1, value=col_a)
        ws1.cell(row=row_idx, column=2, value=col_b)

    # --- Sheet2: 大学別年度変移 ---
    ws2 = wb.create_sheet("大学別年度変移")

    # A列ラベルは年度取得後に書き込む（年度が判明してから更新）
    # 初期値として収入帯ラベルのみ先に設定（年度部分は後で上書き）
    income_band_labels = [
        "平均",           # 行2-4
        "$0-$30,000",     # 行5-7
        "$30,001-$48,000",# 行8-10
        "$48,001-$75,000",# 行11-13
        "$75,001-$110,000",# 行14-16
        "$110,001+",      # 行17-19
    ]

    ws2.cell(row=1, column=1, value="大学名 →").alignment = center
    ws2.column_dimensions["A"].width = px_to_col_width(140)

    # 年度はスクレイピング後に上書きするため、初期は空白
    for row_idx in range(2, 20):
        ws2.column_dimensions["A"].width = px_to_col_width(140)

    wb.save(filepath)
    print(f"✅ Excelファイル作成: {filepath}")
    print(f"   Sheet1「年度最新版」・Sheet2「大学別年度変移」を作成しました")

def create_need_based_excel():
    today = datetime.now().strftime("%Y_%m_%d")
    base_name = f"need-based_{today}"
    script_dir = os.path.dirname(os.path.abspath(__file__))

    college_list_path = os.path.join(script_dir, "College_list.xlsx")
    if not os.path.exists(college_list_path):
        print("❌ College_listがありません")
        return None, None

    college_wb = load_workbook(college_list_path, data_only=True)
    college_ws = college_wb.active
    college_data = []
    for row in college_ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if any(cell is not None for cell in row):
            college_data.append(row)

    print(f"📋 College_listから {len(college_data)} 校を読み込みました")

    filepath = get_unique_filename(script_dir, base_name)
    create_workbook(college_data, filepath)
    return college_data, filepath

def scrape_net_price(driver, college_name):
    """
    戻り値: {
        "years": ["2021-2022", "2022-2023", "2023-2024"],  # 古い順
        "average": ["$19,491", "$17,900", "$17,525"],
        "$0-$30,000": [...],
        "$30,001-$48,000": [...],
        "$48,001-$75,000": [...],
        "$75,001-$110,000": [...],
        "$110,001+": [...],
    }
    """
    wait = WebDriverWait(driver, 10)

    # ① トップページへ移動・大学名を入力して検索
    driver.get(COLLEGE_NAVIGATOR_URL)
    search_box = wait.until(EC.presence_of_element_located((
        By.ID, "ctl00_cphCollegeNavBody_ucSearchMain_txtName"
    )))
    search_box.clear()
    search_box.send_keys(college_name)
    search_btn = wait.until(EC.element_to_be_clickable((
        By.ID, "ctl00_cphCollegeNavBody_ucSearchMain_btnSearch"
    )))
    search_btn.click()
    time.sleep(2)
    print(f"  [{college_name}] 検索完了 → 結果ページ確認中...")

    # ② 検索結果から完全一致の大学リンクをクリック
    try:
        result_links = driver.find_elements(By.CSS_SELECTOR, "table.resultsTable td a strong")
        matched_link = None
        for link in result_links:
            if link.text.strip().lower() == college_name.strip().lower():
                matched_link = link
                break

        if matched_link is None:
            print(f"  [{college_name}] ❌ 完全一致の大学が見つかりませんでした → スキップ")
            return None

        matched_link.click()
        time.sleep(2)
        print(f"  [{college_name}] ✅ 大学ページへ遷移成功")

    except Exception as e:
        print(f"  [{college_name}] ❌ 検索結果クリック失敗 ({e}) → スキップ")
        return None

    # ③ 「Net Price」タブをクリック
    try:
        net_price_tab = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//a[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'net price')]"
        )))
        net_price_tab.click()
        time.sleep(2)
        print(f"  [{college_name}] ✅ Net Priceタブをクリック成功")
    except TimeoutException:
        print(f"  [{college_name}] ⚠️ Net Priceタブが見つかりません。現在のページでデータ取得を試みます...")

    # ④ HTMLをBeautifulSoupで解析・全年度分を取得
    soup = BeautifulSoup(driver.page_source, "html.parser")

    result = {
        "years": [],
        "average": [],
        "$0-$30,000": [],
        "$30,001-$48,000": [],
        "$48,001-$75,000": [],
        "$75,001-$110,000": [],
        "$110,001+": [],
    }

    # College NavigatorのHTMLキーと結果dictキーのマッピング
    income_map = {
        "$0":       "$0-$30,000",
        "$30,001":  "$30,001-$48,000",
        "$48,001":  "$48,001-$75,000",
        "$75,001":  "$75,001-$110,000",
        "$110,001": "$110,001+",
    }

    found_section = False
    for div in soup.find_all("div", class_="tabconstraint"):
        title = div.find("div", class_="tablenames")
        if title and "Average Net Price" in title.text:
            found_section = True
            print(f"  [{college_name}] ✅ 'Average Net Price' セクション発見")
            tables = div.find_all("table", class_="tabular")
            for table in tables:
                # 年度ヘッダーを取得
                thead = table.find("thead")
                if thead:
                    ths = thead.find_all("th")
                    year_headers = [th.get_text(strip=True) for th in ths if th.get_text(strip=True)]
                    if year_headers and not result["years"]:
                        result["years"] = year_headers  # 古い順で格納
                        print(f"  [{college_name}] 年度: {year_headers}")

                # データ行を取得
                for row in table.find_all("tr"):
                    cells = row.find_all("td")
                    if len(cells) >= 2:
                        label = cells[0].get_text(strip=True)
                        values = [c.get_text(strip=True) for c in cells[1:]]

                        if "average net price" in label.lower():
                            result["average"] = values
                        else:
                            for key_prefix, dict_key in income_map.items():
                                if label.replace(" ", "").startswith(key_prefix.replace(" ", "")):
                                    result[dict_key] = values
                                    break
            break

    if not found_section or not result["years"]:
        print(f"  [{college_name}] ❌ Net Priceデータが見つかりませんでした → スキップ")
        return None

    print(f"  [{college_name}] 取得データ概要:")
    print(f"    平均: {result['average']}")
    print(f"    $0-$30,000: {result['$0-$30,000']}")
    return result


if __name__ == "__main__":
    college_data, filepath = create_need_based_excel()
    if not college_data:
        exit()

    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=options)

    try:
        print(f"\n🌐 College Navigatorを開きます: {COLLEGE_NAVIGATOR_URL}")

        wb = load_workbook(filepath)
        ws1 = wb["年度最新版"]
        ws2 = wb["大学別年度変移"]

        failed_colleges = []  # 読み取り失敗した大学リスト

        for col_idx, (college_name, college_type) in enumerate(college_data, start=2):
            if not college_name:
                continue
            print(f"\n{'='*60}")
            print(f"処理中 ({col_idx-1}/{len(college_data)}): {college_name}")
            print(f"{'='*60}")

            try:
                data = scrape_net_price(driver, college_name)
                if data is None:
                    print(f"  [{college_name}] ❌ Excelへの書き込みをスキップ → 失敗リストに追加")
                    failed_colleges.append(college_name)
                    continue

                row_idx = col_idx  # Sheet1の行番号はcol_idxと同じ

                # --- Sheet1「年度最新版」: 最新年度（右端）の値を書き込む ---
                def write_latest(col, values, label):
                    if values:
                        ws1.cell(row=row_idx, column=col, value=values[-1])
                        print(f"  [Sheet1] {label} → {values[-1]}")
                    else:
                        print(f"  [Sheet1] {label} → データなし")

                write_latest(3, data["average"],          "C列（平均）")
                write_latest(4, data["$0-$30,000"],       "D列（$0-$30,000）")
                write_latest(5, data["$30,001-$48,000"],  "E列（$30,001-$48,000）")
                write_latest(6, data["$48,001-$75,000"],  "F列（$48,001-$75,000）")
                write_latest(7, data["$75,001-$110,000"], "G列（$75,001-$110,000）")
                write_latest(8, data["$110,001+"],        "H列（$110,001+）")

                # --- Sheet2「大学別年度変移」: 大学名を1行目に、全年度データを縦に書く ---
                ws2.cell(row=1, column=col_idx, value=college_name).alignment = Alignment(horizontal="center")
                ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = px_to_col_width(140)

                # A列: 年度ラベルを動的に書き込む（最初の大学のデータから年度を取得）
                years = data["years"]  # 例: ["2021-2022", "2022-2023", "2023-2024"]
                income_bands = [
                    ("平均",            2),
                    ("$0-$30,000",      5),
                    ("$30,001-$48,000", 8),
                    ("$48,001-$75,000", 11),
                    ("$75,001-$110,000",14),
                    ("$110,001+",       17),
                ]
                for band_label, start_row in income_bands:
                    for i, yr in enumerate(years):
                        a_label = f"{band_label} {yr}"
                        ws2.cell(row=start_row + i, column=1, value=a_label).alignment = Alignment(horizontal="center")

                def write_all_years(start_row, values, label):
                    for i, v in enumerate(values):
                        ws2.cell(row=start_row + i, column=col_idx, value=v)
                    print(f"  [Sheet2] {label} → {values}")

                write_all_years(2,  data["average"],          f"平均（行2-4）{years}")
                write_all_years(5,  data["$0-$30,000"],       "$0-$30,000（行5-7）")
                write_all_years(8,  data["$30,001-$48,000"],  "$30,001-$48,000（行8-10）")
                write_all_years(11, data["$48,001-$75,000"],  "$48,001-$75,000（行11-13）")
                write_all_years(14, data["$75,001-$110,000"], "$75,001-$110,000（行14-16）")
                write_all_years(17, data["$110,001+"],        "$110,001+（行17-19）")

                print(f"  [{college_name}] ✅ Excel書き込み完了")

            except Exception as e:
                print(f"  [{college_name}] ❌ 予期しないエラー ({e}) → 失敗リストに追加")
                failed_colleges.append(college_name)
                continue

        # --- Sheet1 J列: 読み取り失敗した大学を記録 ---
        if failed_colleges:
            ws1.cell(row=1, column=10, value="読み取り失敗").alignment = Alignment(horizontal="center")
            ws1.column_dimensions["J"].width = px_to_col_width(200)
            for i, name in enumerate(failed_colleges, start=2):
                ws1.cell(row=i, column=10, value=name)
                print(f"  [Sheet1 J列] 失敗大学記録: {name}")
            print(f"\n⚠️  読み取り失敗: {len(failed_colleges)}校 → Sheet1 J列に記録しました")
        else:
            print(f"\n✅ 全大学の読み取りに成功しました")

        # --- Sheet2 A列: 読み取り失敗した大学をA22以降に記録 ---
        if failed_colleges:
            ws2.cell(row=21, column=1, value="読み取り失敗大学").alignment = Alignment(horizontal="center")
            ws2.column_dimensions["A"].width = px_to_col_width(140)
            for i, name in enumerate(failed_colleges, start=22):
                ws2.cell(row=i, column=1, value=name)
                print(f"  [Sheet2 A{i}] 失敗大学記録: {name}")

        wb.save(filepath)
        print(f"\n{'='*60}")
        print(f"✅ 全データを保存しました: {filepath}")
        print(f"{'='*60}")

    finally:
        driver.quit()